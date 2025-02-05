import json
from pprint import pprint
from django.core.serializers.json import DjangoJSONEncoder

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.shortcuts import render
from django.views import View
from django.core import serializers
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from . import models
from . import forms
import statistics

from django.urls import reverse_lazy, reverse
from django.shortcuts import redirect

from django.http import HttpResponse
from .resources import MasterNilaiResource, MasterKegiatanResource, IndikatorKegiatanResources

from master_petugas.models import RoleMitra, AlokasiPetugas, AdministrativeModel
from master_survey.models import SurveyModel, SubKegiatanSurvei

import itertools
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from master_petugas import utils

from master_petugas import models as model_petugas
from master_pegawai import models as model_pegawai
from . import helpers

from django.db.models import Avg
from munapps.mixins import RestrictionsAccess, RestrictionsHttpRequestAccess
from django.db.models.functions import Length

class PenilaianPetugasClassView(LoginRequiredMixin, RestrictionsAccess, View):
    
    def get(self, request):
        context = {
            'title' : 'Kegiatan Penilaian',
            'data_survei' : SurveyModel.objects.filter(~Q(state = 0)),
            'data_subkegiatan' : SubKegiatanSurvei.objects.filter(~Q(status = 0)),
            'data' : models.KegiatanPenilaianModel.objects.all(),
            'roles' : RoleMitra.objects.all(),
            'form' : forms.KegiatanPenilaianForm()
            }
    
        return render(request, 'master_penilaian/kegiatan_penilaian.html', context)


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.KegiatanPenilaianForm(request.POST)
            if form.is_valid():
                safe_check = models.KegiatanPenilaianModel.objects.filter(kegiatan_survey = form.data['kegiatan_survey'])
                if safe_check.exists():
                    return JsonResponse({'status': 'failed', 'message': 'Anda tidak dapat menambahkan penilaian untuk kegiatan survei yang sama.'}, status=200)
                
                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({'status': 'success', 'instance': user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class MasterPenilaianDeleteView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            id = request.POST.get('id')
            penilaian = models.KegiatanPenilaianModel.objects.filter(pk = id)
            if penilaian.exists():
                check_db = models.MasterNilaiPetugas.objects.filter(indikator_penilaian__kegiatan_penilaian = id)
                if check_db.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data sedang digunakan pada master data penilaian'}, status=200)
        
                penilaian.delete()
                return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
            else:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterPenilaianDetailView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            id = request.POST.get('id')
            kegiatan_penilaian = get_object_or_404(models.KegiatanPenilaianModel, pk=id)
            context_data = {
                'id' : kegiatan_penilaian.id,
                'kegiatan_survey' : kegiatan_penilaian.kegiatan_survey.id,
                'tgl_penilaian' : kegiatan_penilaian.tgl_penilaian,
                'status' : kegiatan_penilaian.status,
                'role_permitted' : list(kegiatan_penilaian.role_permitted.values_list('id', flat=True)),
                'role_penilai_permitted' : list(kegiatan_penilaian.role_penilai_permitted.values_list('id', flat=True))
            }

            return JsonResponse({'status' : 'success', 'instance': context_data}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
    

class MasterPenilaianUpdateView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            data = get_object_or_404(models.KegiatanPenilaianModel, pk=request.POST.get('id'))
            form = forms.KegiatanPenilaianForm(request.POST, instance=data)
            if form.is_valid():
                check_db = models.KegiatanPenilaianModel.objects.filter(~Q(id = request.POST.get('id')) & Q(kegiatan_survey = form.cleaned_data['kegiatan_survey']))
                check_db2 = models.MasterNilaiPetugas.objects.filter(Q(indikator_penilaian__kegiatan_penilaian = request.POST.get('id')) | Q(indikator_penilaian__kegiatan_penilaian__kegiatan_survey = form.cleaned_data['kegiatan_survey']))
                if check_db.exists() or check_db2.exists():
                    if 'kegiatan_survey' in form.changed_data:
                        return JsonResponse({'status' : 'failed', 'message': 'Data telah terdaftar atau sedang digunakan pada master data penilaian.'}, status=200)
                
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class PenilaianGetBySurveiClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            data = models.KegiatanPenilaianModel.objects.filter(kegiatan_survey__survey = request.POST.get('survey_id')).values('id', 'kegiatan_survey__nama_kegiatan')
            if data.exists():
                return JsonResponse({"status": "success", "instance": list(data)}, status=200)

            return JsonResponse({"status": "failed", "message" : "Data tidak tersedia", "instance" : []}, status=200)

        return JsonResponse({"error": ""}, status=400)
    

class AlokasiGetBySurveiClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            penilaian_id = int(request.POST.get('penilaian_id'))
            kegiatan_penilaian = get_object_or_404(models.KegiatanPenilaianModel, pk=penilaian_id)
            
            qry = AlokasiPetugas.objects.filter(sub_kegiatan__survey=request.POST.get('survey_id'))
            penilai = qry.filter(petugas = None).values('id', 'pegawai__nip', 'pegawai__name', 'role', 'role__jabatan' ).order_by('pegawai__name', 'role__jabatan')

            data = qry.filter(pegawai = None).values('id', 'petugas__kode_petugas', 'petugas__nama_petugas', 'role', 'role__jabatan' ).order_by('petugas__nama_petugas', 'role__jabatan')
            role_permitted = kegiatan_penilaian.role_permitted.values_list('id', flat=True)
            role_permitted_penilai = kegiatan_penilaian.role_penilai_permitted.values_list('id', flat=True)
            petugas_data, penilai_data = [], []

            for dt in data :
                if dt['role'] in role_permitted:
                    petugas_data.append(dt)
            
            for dt in penilai :
                if dt['role'] in role_permitted_penilai:
                    penilai_data.append(dt)
         
            return JsonResponse({"instance": list(petugas_data), "instance_2": list(penilai)}, status=200)
        
        return JsonResponse({"error": ""}, status=400)


class MasterPenilaianJsonResponseClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):    
        data = self._datatables(request)
        return HttpResponse(json.dumps(data, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.KegiatanPenilaianModel.objects

        if datatables.get('survei_filter'):
            data = data.filter(kegiatan_survey__survey = datatables.get('survei_filter'))
        
        if datatables.get('status_filter'):
            data = data.filter(status = datatables.get('status_filter'))

        if datatables.get('role_filter'):
            data = data.filter(role_permitted = datatables.get('role_filter'))

        if search:
            data = data.filter(
                Q(kegiatan_survey__nama_kegiatan__icontains=search)|
                Q(kegiatan_survey__survey__nama__icontains=search)|
                Q(tgl_penilaian__icontains=search)|
                Q(status__icontains=search)|
                Q(role_permitted__jabatan__icontains=search)|
                Q(role_penilai_permitted__jabatan__icontains=search)
            )

        data = data.exclude(Q(kegiatan_survey=None)|Q(tgl_penilaian=None)|Q(status=None)|Q(role_permitted=None)|Q(role_penilai_permitted=None)|Q(status=None))
        records_total = data.count()
        records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list
        
        data = [
           
            {
                'kegiatan_survey__nama_kegiatan': obj.kegiatan_survey.nama_kegiatan,
                'tgl_penilaian': obj.tgl_penilaian.strftime('%d-%b-%Y'),
                'role_permitted__jabatan' : ', '.join(obj.role_permitted.values_list('jabatan', flat=True)),
                'role_penilai_permitted__jabatan' : ', '.join(obj.role_penilai_permitted.values_list('jabatan', flat=True)),
                'status': f'<span class="badge badge-primary-lighten p-1"> {obj.get_status_display()} </span>'  if obj.status == '0' else f'<span class="badge badge-primary-lighten p-1"> {obj.get_status_display()} </span>',
                'aksi': f'<a href="javascript:void(0);" onclick="editKegiatanPenilaian({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="hapusKegiatanPenilaian({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>'
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class MasterPenilaianExportView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request):
        resource = MasterKegiatanResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Kegiatan Penilaian.xls"'
        return response 


#INDIKATOR PENILAIAN
class IndikatorPenilaianClassView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request):
        context = {
            'title' : 'Indikator Penilaian',
            'form' : forms.IndikatorPenilaianForm()
            }

        return render(request, 'master_penilaian/indikator_penilaian.html', context)


    def post(self, request):

        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.IndikatorPenilaianForm(request.POST)
            if form.is_valid():
                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class IndiakatorPenilaianJsonResponseClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.IndikatorPenilaian.objects
        if search:
            data = data.filter(
                Q(nama_indikator__icontains=search)|
                Q(deskripsi_penilaian__icontains=search)
            )

        data = data.exclude(Q(nama_indikator=None)|Q(deskripsi_penilaian=None))
        records_total = data.count()
        records_filtered = records_total
        
        data = data.order_by(order_col_name)
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'nama_indikator': obj.nama_indikator,
                'deskripsi_penilaian': obj.deskripsi_penilaian,
                'aksi': f'<td><a href="javascript:void(0);" onclick="editIndikator({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="deleteIndikator({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a></td>'
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class IndikatorPenilaianDeleteView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            id = request.POST.get('id')
            indikator_penilaian = models.IndikatorPenilaian.objects.filter(pk = id)
            if indikator_penilaian.exists():
                check_kegiatan_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(indikator_penilaian = id)
                if check_kegiatan_penilaian.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)

                check_db = models.MasterNilaiPetugas.objects.filter(indikator_penilaian__indikator_penilaian = id)
                if check_db.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)
            
                indikator_penilaian.delete()
                return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
            else:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)
    
class IndikatorPenilaianDetailView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            id = request.POST.get('id')
            
            # check_db = models.MasterNilaiPetugas.objects.filter(penilaian__indikator_penilaian = id)
            # if check_db.exists():
                # return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)
            
            indikator_penilaian = models.IndikatorPenilaian.objects.filter(pk=id)
            if indikator_penilaian.exists():
                return JsonResponse({'status' : 'success', 'instance': list(indikator_penilaian.values())[0]}, status=200)
            else:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
    

class IndikatorPenilaianUpdateView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.IndikatorPenilaian, pk=request.POST.get('id'))

            # check_db = models.MasterNilaiPetugas.objects.filter(penilaian__indikator_penilaian = request.POST.get('id'))
            # if check_db.exists():
                # return JsonResponse({'status' : 'failed', 'message': 'Data Indikator Penilaian telah digunakan pada master data penilaian.'}, status=200)

            form = forms.IndikatorPenilaianForm(request.POST, instance=data)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


# INDIKATOR KEGIATAN PENILAIAN
class IndikatorKegiatanPenilaianClassView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request):
        context = {
            'title' : 'Kegiatan Penilaian',
            'data_kegiatan': models.KegiatanPenilaianModel.objects.all(),
            'data_indikator': models.IndikatorPenilaian.objects.all(),
            'form' : forms.IndikatorKegiatanPenilaianForm()
            }
        
        return render(request, 'master_penilaian/indikator_kegiatan_penilaian.html', context)


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.IndikatorKegiatanPenilaianForm(request.POST)
            if form.is_valid():
                check_db2 = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = form.data['kegiatan_penilaian'], indikator_penilaian = form.data['indikator_penilaian'])
                check_db3 = models.MasterNilaiPetugas.objects.filter(indikator_penilaian__kegiatan_penilaian = form.data['kegiatan_penilaian'], indikator_penilaian__indikator_penilaian = form.data['indikator_penilaian'])
                if check_db2.exists() or check_db3.exists():
                    return JsonResponse({'status' : 'failed', 'message': f'Indikator Penilaian telah terdaftar untuk  pada database.'}, status=200)
                else:
                    instance = form.save()
                    user_instance = serializers.serialize('json', [ instance, ])
                    return JsonResponse({'status' : 'success', "instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({'status' : 'failed', "error": form.errors}, status=400)
            
        return JsonResponse({'status' : 'failed', "error": ""}, status=400)
    
class IndikatorKegiatanPenilaianJsonResponseClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.IndikatorKegiatanPenilaian.objects

        if datatables.get('kegiatan_filter'):
            data = data.filter(kegiatan_penilaian = datatables.get('kegiatan_filter'))
        
        if datatables.get('indikator_filter'):
            data = data.filter(indikator_penilaian = datatables.get('indikator_filter'))
        
        if search:
            data = data.filter(
                Q(kegiatan_penilaian__kegiatan_survey__survey__nama=search)|
                Q(kegiatan_penilaian__kegiatan_survey__nama_kegiatan__icontains=search)|
                Q(indikator_penilaian__nama_indikator__icontains=search)
            )

        data = data.exclude(Q(kegiatan_penilaian=None)|Q(indikator_penilaian=None))
        records_total = data.count()
        records_filtered = records_total
        
        data = data.order_by(order_col_name)
        
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'kegiatan_penilaian__kegiatan_survey__nama_kegiatan': obj.kegiatan_penilaian.kegiatan_survey.nama_kegiatan,
                'indikator_penilaian__nama_indikator': obj.indikator_penilaian.nama_indikator,
                'scale': f'{obj.get_scale_display()} [Min: <b>{obj.n_min}</b>, Max: <b>{obj.n_max}</b>]',
                'aksi': f'<a href="javascript:void(0);" onclick="editIndikatorKegiatan({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="deleteIndikatorKegiatan({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>'
            } for obj in object_list
        ]

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class IndikatorKegiatanPenilaianDeleteView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            id = request.POST.get('id')
            indikator_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(pk = id)
            if indikator_penilaian.exists():
                check_db = models.MasterNilaiPetugas.objects.filter(indikator_penilaian = id)
                if check_db.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Indikator Penilaian sedang digunakan pada master data penilaian.'}, status=200)

                indikator_penilaian.delete()
                return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
            else:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)
    
class IndikatorKegiatanPenilaianDetailView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            id = request.POST.get('id')

            indikator_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(pk=id)

            if indikator_penilaian.exists():
                check_db = models.MasterNilaiPetugas.objects.filter(penilaian = id)
                if check_db.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada master data penilaian.'}, status=200)
            
                return JsonResponse({'status' : 'success', 'instance': list(indikator_penilaian.values())[0]}, status=200)
            else:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 
    
class IndikatorKegiatanPenilaianUpdateView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            data = get_object_or_404(models.IndikatorKegiatanPenilaian, pk=request.POST.get('id'))
            form = forms.IndikatorKegiatanPenilaianForm(request.POST, instance=data)
            if form.is_valid():
                check_db = models.MasterNilaiPetugas.objects.filter(indikator_penilaian = form.data['id'])
                if check_db.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Indikator penilaian sedang digunakan pada master data penilaian.'}, status=200)
                
                check_db2 = models.IndikatorKegiatanPenilaian.objects.filter(~Q(id = form.data['id'] ) & Q(kegiatan_penilaian = form.data['kegiatan_penilaian'] ) & Q(indikator_penilaian = form.data['indikator_penilaian']) )
                if check_db2.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Indikator penilaian telah terdaftar pada database.'}, status=200)
                
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({'status' : 'success', "instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({'status' : 'failed', "error": form.errors}, status=400)
        return JsonResponse({'status' : 'failed', "error": ""}, status=400)

class IndikatorKegiatanPenilaianExportView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request):
    
        resource = IndikatorKegiatanResources()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Indikator Kegiatan Penilaian.xls"'
        return response
    

# NILAI MITRA
class NilaiMitraClassView(LoginRequiredMixin, RestrictionsAccess, View):
    
    def get(self, request):
        context = {
            'title' : 'Nilai Mitra',
            'data_mitra' : AlokasiPetugas.objects.filter(pegawai = None).all().order_by('petugas__nama_petugas'),
            'adm_prov' : AdministrativeModel.objects.annotate(text_len=Length('code')).filter(text_len=2).order_by('region'),
            'data_role' : RoleMitra.objects.all().order_by('jabatan'),
            'data_survei' : SurveyModel.objects.all().order_by('nama'),
            'data_kegiatan' : SubKegiatanSurvei.objects.all().order_by('nama_kegiatan'),
            'data_indikator_penilaian' : models.IndikatorPenilaian.objects.all(),
            'data_kegiatan_penilaian' : models.KegiatanPenilaianModel.objects.filter(status = '1').order_by('-tgl_penilaian'),
            'data_pegawai' : model_pegawai.MasterPegawaiModel.objects.all().order_by('name'),
            'form' : forms.PenilaianMitraForm(),
            'form_upload' : forms.NilaiFormUpload()
            }

        return render(request, 'master_penilaian/nilai-mitra.html', context)

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.PenilaianMitraForm(request.POST)
            if form.is_valid():
                df = form.cleaned_data
                objs_create, objs_update = 0, 0
                for idx, dt in enumerate(df):
                    modelPenilaian = models.MasterPenilaianPetugas
                    modelNilai = models.MasterNilaiPetugas
                    petugas = AlokasiPetugas.objects.get(pk = dt['petugas'])
                    penilai = AlokasiPetugas.objects.get(pk = dt['penilai'])
                    indikator_penilaian = models.IndikatorKegiatanPenilaian.objects.get(pk = dt['penilaian'])

                    role_permitted = list(indikator_penilaian.kegiatan_penilaian.role_permitted.values_list('id', flat=True))
                    role_permitted_penilai = list(indikator_penilaian.kegiatan_penilaian.role_penilai_permitted.values_list('id', flat=True))
                    
                    if petugas.role.id not in role_permitted:
                        return JsonResponse({"status": "failed", "messages":  f'Data alokasi petugas dengan role {petugas.role.jabatan} tidak diizinkan untuk mengikuti kegiatan penilaian "<i>{indikator_penilaian.kegiatan_penilaian.nama_kegiatan}</i>"'})
                    
                    if penilai.role.id not in role_permitted_penilai:
                        return JsonResponse({"status": "failed", "messages":  f'Data alokasi pegawai dengan role {penilai.role.jabatan} tidak diizinkan untuk mengikuti kegiatan penilaian "<i>{indikator_penilaian.kegiatan_penilaian.nama_kegiatan}</i>"'})
            
                    nilai = dt['nilai']
                    catatan = dt['catatan']

                    db_check = modelPenilaian.objects.filter(petugas = petugas, penilai = penilai, detail_nilai__indikator_penilaian__kegiatan_penilaian = indikator_penilaian.kegiatan_penilaian).values('id', 'petugas', 'penilai').distinct()
                    if db_check.exists():
                        db_check2 = models.MasterNilaiPetugas.objects.filter(penilaian = db_check.first()['id'], indikator_penilaian = indikator_penilaian.pk)
                        if db_check2.exists():
                            nilai_mitra_update = db_check2.first()
                            nilai_mitra_update.nilai = nilai
                            nilai_mitra_update.catatan = catatan
                            nilai_mitra_update.save()
                            objs_update += 1
                        else:
                            modelNilai(
                                penilaian = modelPenilaian.objects.get(pk = db_check.first()['id']),
                                indikator_penilaian = indikator_penilaian,
                                nilai = nilai,
                                catatan = catatan,
                            ).save()
                            objs_create += 1
                    else:
                        row_affected = modelPenilaian.objects.create(petugas = petugas, penilai = penilai, state = '2')
                        modelNilai(
                            penilaian = row_affected,
                            indikator_penilaian = indikator_penilaian,
                            nilai = nilai,
                            catatan = catatan,
                        ).save()
                        objs_create += 1

                msg = ''
                if objs_create > 0:
                    msg += f"Data <strong>berhasil</strong> ditambahkan.<br>"
                if objs_update > 0:
                    msg += f"Data <strong>berhasil</strong> diperbarui.<br>"

                return JsonResponse({"status": "success", "messages":  msg})
            
            return JsonResponse({"error": form.errors}, status=400)
        else:
            return JsonResponse({"error": ""}, status=400)

class NilaiMitraJsonResponseClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.MasterPenilaianPetugas.objects

        if datatables.get('region_code'):
            data = data.filter(petugas__petugas__adm_id__code__icontains = datatables.get('region_code'))
            
        if datatables.get('mitra_filter'):
            data = data.filter(petugas__petugas = datatables.get('mitra_filter'))
            
        if datatables.get('pegawai_filter'):
            data = data.filter(penilai__pegawai = datatables.get('pegawai_filter'))
        
        if datatables.get('role_filter'):
            data = data.filter(petugas__role = datatables.get('role_filter'))

        if datatables.get('survei_filter'):
            data = data.filter(detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey__survey = datatables.get('survei_filter'))

        if datatables.get('kegiatan_filter'):
            data = data.filter(detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey = datatables.get('kegiatan_filter'))
    
        if search:
            data = data.filter(
                Q(penilai__pegawai__name__icontains=search)|
                Q(penilai__pegawai__nip__icontains=search)|
                Q(petugas__petugas__kode_petugas__icontains=search)|
                Q(petugas__petugas__adm_id__region__icontains=search)|
                Q(petugas__petugas__adm_id__code__icontains=search)|
                Q(petugas__petugas__nama_petugas__icontains=search)|
                Q(petugas__role__jabatan__icontains=search)|
                Q(penilai__role__jabatan__icontains=search)|
                Q(detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey__nama_kegiatan__icontains=search)|
                Q(detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator__icontains=search)|
                Q(detail_nilai__nilai__icontains=search)|
                Q(detail_nilai__catatan__icontains=search)
            )

        data = data.order_by(order_col_name).values('penilai__pegawai__name', 'penilai__pegawai__nip', 'petugas__petugas__id','petugas__petugas__adm_id__region', 'petugas__petugas__kode_petugas', 'petugas__petugas__nama_petugas', 'petugas__role__jabatan', 'penilai__role__jabatan','detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey__nama_kegiatan', 'detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator', 'detail_nilai__nilai', 'detail_nilai__catatan')
        records_total = data.count()
        records_filtered = records_total

        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list
        
        data = []
        for obj in object_list:
            # kec = AdministrativeModel.objects.filter(code = obj['petugas__petugas__adm_id__code'][:-3]).first()
            data.append({
                'petugas__petugas__adm_id__region': obj['petugas__petugas__adm_id__region'],
                'petugas__petugas__kode_petugas': obj['petugas__petugas__kode_petugas'],
                'petugas__petugas__nama_petugas': f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": obj["petugas__petugas__id"]})}" class="text-body" target="_blank">{obj["petugas__petugas__nama_petugas"]}</a>',
                'petugas__role__jabatan': obj["petugas__role__jabatan"],
                'detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey__nama_kegiatan': obj["detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey__nama_kegiatan"],
                'detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator': obj["detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator"],
                'detail_nilai__nilai': obj["detail_nilai__nilai"],
                'detail_nilai__catatan' :  obj["detail_nilai__catatan"],
                'penilai__pegawai__name' :  f'{obj["penilai__pegawai__name"]} ({obj["penilai__role__jabatan"]})',
            })
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class NilaiMitraDetailClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            id = request.POST.get('id')

            nilai_mitra = models.MasterNilaiPetugas.objects.filter(pk=id)

            if nilai_mitra.exists():
                return JsonResponse({'status' : 'success', 'instance': list(nilai_mitra.values())[0]}, status=200)
            else:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class NilaiMitraUpdateClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.MasterNilaiPetugas, pk=request.POST.get('id'))
            
            form = forms.PenilaianMitraForm(request.POST, instance=data)
            if form.is_valid():

                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)

class NilaiMitraDeleteClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            penilaian_mitra = models.MasterPenilaianPetugas.objects.filter(id = request.POST.get('id'))
            if penilaian_mitra.exists() == False:
                return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
            penilaian_mitra.first().delete()
            return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)
    
class NilaiMitraExportClassView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request):
        resource = MasterNilaiResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Master Nilai Mitra.xls"'
        return response 
    
class NilaiMitraTemplateClassView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request, *args, **kwargs): 
        kegiatan_penilaian = request.GET.get('kegiatan')
        penilai = request.GET.get('penilai')

        kegiatan = models.KegiatanPenilaianModel.objects.filter(pk=kegiatan_penilaian)
        penilai = AlokasiPetugas.objects.filter(pk = penilai)

        if not kegiatan.exists() or not penilai.exists():
            return redirect(reverse('master_penilaian:nilai-mitra'))
        
        kegiatan = kegiatan.first()
        penilai = penilai.first()

        role_permitted = list(kegiatan.role_permitted.values_list('id', flat=True))
        role_penilai__permitted = list(kegiatan.role_penilai_permitted.values_list('id', flat=True))
        if penilai.role.pk not in role_penilai__permitted:
            return redirect(reverse('master_penilaian:nilai-mitra'))

        alokasi_petugas = AlokasiPetugas.objects.filter(sub_kegiatan = kegiatan.kegiatan_survey, pegawai = None)
        indikator_kegiatan_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = kegiatan.id)
        indikator = list(indikator_kegiatan_penilaian.values_list('indikator_penilaian__nama_indikator', flat=True))
        catatan_indikator = [f"Catatan Personal {dt}" for dt in indikator]
        indikator_penilaian = [f"Penilaian Indikator {dt}" for dt in indikator]
        headers = ['No', 'ID Alokasi Mitra', 'ID Kegiatan Penilaian', 'ID Penilai', 'Kode Petugas', 'Nama Petugas', 'Survei', 'Jabatan', 'Kegiatan Penilaian', 'Nama Penilai (Organik)'] + indikator_penilaian + catatan_indikator

        head_row = 2
        alphabet_string = utils.generate_headers_excel(100)
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Upload Nilai Mitra Petugas'
        header_cols = np.array(ws[f'A{head_row}':f'{alphabet_string[len(headers)]}{head_row}'])
        
        for v,c in zip(headers, header_cols.T.flatten()):
            c.font = Font(name='Calibri', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type = "solid")
            c.value = v

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50
    
        ws.merge_cells(f'A1:{alphabet_string[10:10 + len(indikator_penilaian)*2][-1]}1')
        ws['A1'] = 'Template Upload Nilai Mitra'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Calibri',bold=True, size=14)

        no_ = 0
        for idx, dt_alokasi in enumerate(alokasi_petugas):
            if dt_alokasi.role.pk not in role_permitted:
                continue
        
            db_query_check = models.MasterPenilaianPetugas.objects.filter(petugas = dt_alokasi.id, penilai = penilai.pk)
            if db_query_check.exists():
                db_query_check2 = db_query_check.filter(detail_nilai__indikator_penilaian__kegiatan_penilaian = kegiatan.id).values('detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator', 'detail_nilai__nilai', 'detail_nilai__catatan')
                if db_query_check2.exists():
                    ws[f'A{no_+3}'] = no_ + 1
                    ws[f'B{no_+3}'] = dt_alokasi.id
                    ws[f'C{no_+3}'] = kegiatan.id
                    ws[f'D{no_+3}'] = penilai.id
                    ws[f'E{no_+3}'] = dt_alokasi.petugas.kode_petugas
                    ws[f'F{no_+3}'] = dt_alokasi.petugas.nama_petugas
                    ws[f'G{no_+3}'] = dt_alokasi.sub_kegiatan.survey.nama
                    ws[f'H{no_+3}'] = dt_alokasi.role.jabatan
                    ws[f'I{no_+3}'] = kegiatan.kegiatan_survey.nama_kegiatan
                    ws[f'J{no_+3}'] = penilai.pegawai.name

                    for c in alphabet_string[:10]:
                        ws[f'{c}{no_+3}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")

                    for idx2, db_row in enumerate(db_query_check2):
                        db_row_indikator = db_row['detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator']
                        indikator_index = indikator.index(db_row_indikator) if db_row_indikator in indikator else -1

                        if indikator_index >= 0:
                            ws[f'{alphabet_string[10 + indikator_index]}{no_+3}'] = db_row['detail_nilai__nilai']
                            ws[f'{alphabet_string[10 + len(indikator) + indikator_index]}{no_+3}'] = db_row['detail_nilai__catatan']
                            ws[f'{alphabet_string[10 + indikator_index]}{no_+3}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")
                            ws[f'{alphabet_string[10 + len(indikator) + indikator_index]}{no_+3}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")
                    no_ += 1
            else:
                ws[f'A{no_+3}'] = no_ + 1
                ws[f'B{no_+3}'] = dt_alokasi.id
                ws[f'C{no_+3}'] = kegiatan.id
                ws[f'D{no_+3}'] = penilai.id
                ws[f'E{no_+3}'] = dt_alokasi.petugas.kode_petugas
                ws[f'F{no_+3}'] = dt_alokasi.petugas.nama_petugas
                ws[f'G{no_+3}'] = dt_alokasi.sub_kegiatan.survey.nama
                ws[f'H{no_+3}'] = dt_alokasi.role.jabatan
                ws[f'I{no_+3}'] = kegiatan.kegiatan_survey.nama_kegiatan
                ws[f'J{no_+3}'] = penilai.pegawai.name
                no_ += 1

        ws.column_dimensions['B'].hidden= True
        ws.column_dimensions['C'].hidden= True
        ws.column_dimensions['D'].hidden= True
        # ws[f'A{no_+4}'] = 'Keterangan:'
        # ws[f'D{no_+5}'] = 'Telah tercatat pada database, ubah data untuk mengupdate nilai mitra.'
        # ws[f'A{no_+5}'].fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type = "solid")
        # ws[f'A{no_+4}'].font = Font(name='Calibri', size=10)
        # ws[f'D{no_+5}'].font = Font(name='Calibri', size=10)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Nilai Mitra.xlsx'

        wb.save(response)
        return response

class NilaiMitraUploadClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            form = forms.NilaiFormUpload(request.POST, request.FILES)
            model = models.MasterPenilaianPetugas
            modelNilai = models.MasterNilaiPetugas  

            if form.is_valid():
                df = form.cleaned_data
                objs_create = 0
                objs_update = 0
                for dt in df:
                    petugas = dt['petugas']
                    penilai = dt['penilai']
                    penilaian = dt['penilaian']
                    nilai = dt['nilai']
                    catatan = dt['catatan']
                    
                    db_check = model.objects.filter(petugas = petugas.pk, penilai = penilai.pk, detail_nilai__indikator_penilaian = penilaian.pk).values('id', 'petugas', 'penilai').distinct()
                    if db_check.exists():
                        db_check2 = models.MasterNilaiPetugas.objects.filter(penilaian = db_check.first()['id'], indikator_penilaian = penilaian.pk)
                        if db_check2.exists():
                            nilai_mitra_update = db_check2.first()
                            nilai_mitra_update.nilai = nilai
                            nilai_mitra_update.catatan = catatan
                            nilai_mitra_update.save()
                            objs_update += 1
                        else:
                            modelNilai(
                                penilaian = db_check.first(),
                                indikator_penilaian = penilaian,
                                nilai = nilai,
                                catatan = catatan,
                            ).save()
                            objs_create += 1
                    else:
                        row_affected = model.objects.create(petugas = petugas, penilai = penilai, state = '2')
                        modelNilai(
                            penilaian = row_affected,
                            indikator_penilaian = penilaian,
                            nilai = nilai,
                            catatan = catatan,
                        ).save()
                        objs_create += 1

                msg = ''
                if objs_create > 0:
                    msg += f"Data <strong>berhasil</strong> ditambahkan.<br>"
                if objs_update > 0:
                    msg += f"Data <strong>berhasil</strong> diperbarui.<br>"

                return JsonResponse({"status": "success", "messages":  msg})
            else:
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})
            
        return JsonResponse({"error": ""}, status=403)  

class GenerateTableNilaiClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            kegiatan_penilaian_id = request.POST.get('kegiatan_penilaian')
            kegiatan = models.KegiatanPenilaianModel.objects.filter(pk=kegiatan_penilaian_id)
            if kegiatan.exists() == False:
                return JsonResponse({'status': 'failed', 'data': 'Data tidak tersedia'})
            
            kegiatan = kegiatan.first()
            kegiatan_survey_id = kegiatan.kegiatan_survey.id

            if request.POST.get('filter_role_nilai_mitra'):
                alokasi_petugas = AlokasiPetugas.objects.filter(sub_kegiatan = kegiatan_survey_id, role = request.POST.get('filter_role_nilai_mitra'), pegawai = None)
            else:
                alokasi_petugas = AlokasiPetugas.objects.filter(sub_kegiatan = kegiatan_survey_id, pegawai = None)

            master_nilai = models.MasterNilaiPetugas.objects.filter(indikator_penilaian__kegiatan_penilaian = kegiatan_penilaian_id)
            indikator_penilaian = list(master_nilai.values_list('indikator_penilaian__indikator_penilaian__nama_indikator', flat=True).distinct())
            indikator_col = [ f"Penilaian <br> {indikator}" for indikator in indikator_penilaian]
            catatan_col = [ f"Catatan {indikator}" for indikator in indikator_penilaian]
            
            headers = ['ID Mitra', 'Kec.',	'Mitra',	'Jabatan', 'Kegiatan Penilaian']
            headers += indikator_col + ['Rerata', 'Catatan', 'Penilai'] + ['Aksi']
            
            thead = '<tr>'
            for header in headers:
                if 'Penilaian' in header:
                    thead += f'<th class="text-center">{header}</th>'
                else:
                    thead += f'<th>{header}</th>'
            thead += '</tr>'

            tbody_data = []
            # print('Total petugas dinilai: ', alokasi_petugas)
            for idx, dt_alokasi in enumerate(alokasi_petugas):
                kec = AdministrativeModel.objects.filter(code = dt_alokasi.petugas.adm_id.code[:-3]).first()
                dt_row = []
                dt_row.append(dt_alokasi.petugas.id) #[1]
                dt_row.append(dt_alokasi.petugas.kode_petugas) #[2]
                dt_row.append(f'Kec. {kec.region}') #[3]
                dt_row.append(dt_alokasi.petugas.nama_petugas) #[4]
                dt_row.append(dt_alokasi.role.jabatan) #[5]
                dt_row.append(kegiatan.kegiatan_survey.nama_kegiatan) #[6]

                db_query_check = models.MasterPenilaianPetugas.objects.filter(petugas = dt_alokasi.id, detail_nilai__indikator_penilaian__kegiatan_penilaian = kegiatan.id).values('penilai', 'penilai__pegawai__name', 'petugas', 'detail_nilai__indikator_penilaian__kegiatan_penilaian').distinct()

                if db_query_check.exists():
                    for idx2, db_row in enumerate(db_query_check.values('id', 'penilai__pegawai__name')): # Looping berdasarkan penilai
                        nilai = [0] * len(indikator_penilaian)
                        catatan = '<ul>'
                        db_query_nilai = models.MasterNilaiPetugas.objects.filter(penilaian = db_row['id']).values('nilai', 'catatan', 'indikator_penilaian__indikator_penilaian__nama_indikator')
                        for db_child in db_query_nilai: # Looping berdasarkan indikator penilaian
                            db_row_indikator = db_child['indikator_penilaian__indikator_penilaian__nama_indikator']
                            indikator_index = indikator_penilaian.index(db_row_indikator) if db_row_indikator in indikator_penilaian else -1
                            if indikator_index >= 0:
                                # nilai.insert(indikator_index + 1, db_child['nilai'])
                                nilai[indikator_index] = db_child['nilai']
                                if len(db_child['catatan']) > 0:
                                    catatan += f"<li>{db_child['catatan']}</li>"

                        catatan += '</ul>'
                        rerata_nilai = round(statistics.mean(nilai), 2)
                        tbody_data.append([db_row['id']] + dt_row + nilai + [rerata_nilai, catatan, db_row['penilai__pegawai__name']])

            tbody = ''
            for data in tbody_data:
                tbody += '<tr>'
                for idx, dt in enumerate(data):
                    if idx == 0:
                        button_action = f'<td class="text-center"><button class="btn btn-primary" onclick="updateNilaiMitra({dt})" >Edit</button> <button class="btn btn-danger" onclick="deleteNilaiMitra({dt})">Hapus</button></td>'
                    elif idx == 1:
                        continue
                    elif idx in [2, 3, 6]:
                        tbody += f'<td>{dt}</td>'
                    elif idx == 4:
                        tbody += f'<td><a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": data[1]})}" class="text-body" target="_blank">{data[4]}</a></td>'
                    else:
                        txt_style = '' if (idx+1) == len(data) or (idx+1) == (len(data) - 1) or (idx+1) == (len(data) - 1) else 'text-center'
                        tbody += f'<td class="{txt_style}">{dt}</td>'

                tbody += button_action
                tbody += '</tr>'

            return JsonResponse({'status' : 'success', 'data': {'thead' : thead, 'tbody' : tbody}}, status=200)

        return JsonResponse({'status': 'Invalid request'}, status=400)

class GetNilaiMitraClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            id_petugas = request.POST.get('id_petugas')
            id_penilai = request.POST.get('id_penilai')
            id_kegiatan = request.POST.get('id_kegiatan')
            id_penilaian = request.POST.get('id_penilaian')
            
            if id_penilaian: # Request from update nilai
                penilaian = models.MasterPenilaianPetugas.objects.filter(pk = id_penilaian)
                if penilaian.exists():
                    nilai_mitra = penilaian.values('petugas', 'penilai', 'state', 'id', 'detail_nilai__indikator_penilaian__n_min', 'detail_nilai__indikator_penilaian__n_max', 'detail_nilai__indikator_penilaian', 'detail_nilai__indikator_penilaian__kegiatan_penilaian', 'detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator', 'detail_nilai__nilai' , 'detail_nilai__catatan')
                    data_kegiatan = models.KegiatanPenilaianModel.objects.filter(pk = nilai_mitra.first()['detail_nilai__indikator_penilaian__kegiatan_penilaian']).values('id', 'kegiatan_survey__nama_kegiatan')

                    list_indicators = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = data_kegiatan.first().get('id'))
                    opt = helpers.convert_table_penilaian(nilai_mitra, list_indicators)
                 
                    data_mitra = AlokasiPetugas.objects.filter(pk=nilai_mitra.first().get('petugas')).values('id', 'petugas__kode_petugas', 'petugas__nama_petugas', 'role__jabatan', 'sub_kegiatan__survey__id', 'sub_kegiatan__survey__nama', 'sub_kegiatan__nama_kegiatan')
                    data_penilai = AlokasiPetugas.objects.filter(pk=nilai_mitra.first().get('penilai')).values('id', 'pegawai__nip', 'pegawai__name', 'role__jabatan')

                    return JsonResponse({'status': 'success', 'data_kegiatan': data_kegiatan.first(), 'data_mitra': data_mitra.first(), 'data_penilai': data_penilai.first(), 'nilai_mitra': list(nilai_mitra), 'opt' : opt}, status=200)
            else:
                data_mitra = AlokasiPetugas.objects.filter(pk=id_petugas).values('id', 'petugas__kode_petugas', 'petugas__nama_petugas', 'role__jabatan', 'sub_kegiatan__survey__id', 'sub_kegiatan__survey__nama', 'sub_kegiatan__nama_kegiatan')
                data_penilai = AlokasiPetugas.objects.filter(pk=id_penilai).values('id', 'pegawai__nip', 'pegawai__name', 'role__jabatan')
                data_kegiatan = models.KegiatanPenilaianModel.objects.filter(pk = id_kegiatan).values('id', 'kegiatan_survey__nama_kegiatan')

                if (data_mitra.exists() and data_penilai.exists() and data_kegiatan.exists()):
                    indikator_penilaian = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = id_kegiatan)
                    nilai_mitra = models.MasterPenilaianPetugas.objects.filter(petugas = id_petugas, penilai = id_penilai, detail_nilai__indikator_penilaian__kegiatan_penilaian=id_kegiatan).values('state', 'id', 'detail_nilai__indikator_penilaian', 'detail_nilai__indikator_penilaian__indikator_penilaian__nama_indikator', 'detail_nilai__nilai' , 'detail_nilai__catatan' )
                    if nilai_mitra.exists() is False:
                        opt = helpers.convert_table_penilaian(nilai_mitra, indikator_penilaian, empty=True)
                    else:
                        opt = helpers.convert_table_penilaian(nilai_mitra, indikator_penilaian)
                    return JsonResponse({'status': 'success', 'data_kegiatan': data_kegiatan.first(), 'data_mitra': data_mitra.first(), 'data_penilai': data_penilai.first(), 'nilai_mitra': list(nilai_mitra), 'opt' : opt}, status=200)
            return JsonResponse({'status': 'failed', 'message': data_kegiatan.first()}, status=200)
        
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class GetPenilaiClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax and request.POST.get('sub_kegiatan') and request.POST.get('kegiatan_penilaian'):
            penilai = AlokasiPetugas.objects.filter(sub_kegiatan = request.POST.get('sub_kegiatan'), petugas=None)
            kegiatan_penilaian = models.KegiatanPenilaianModel.objects.filter(pk = request.POST.get('kegiatan_penilaian'))
            if penilai.exists() and kegiatan_penilaian.exists():
                role_penilai_permitted = list(kegiatan_penilaian.first().role_penilai_permitted.values_list('id', flat=True))
                opt = '<option selected="" value="">-- Pilih Penilai --</option>'
                idx = 1
                for dt in penilai.values('id', 'pegawai__name', 'pegawai__nip', 'role').order_by('pegawai__name'):
                    if dt['role'] in role_penilai_permitted:
                        opt += f'<option value="{dt["id"]}">{idx}. {dt["pegawai__name"]} [{dt["pegawai__nip"]}]</option>'
                        idx += 1
                return JsonResponse({'status': 'success', 'data': opt}, status=200)
            return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia.'}, status=200)
        
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class EntryPenilaianClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        
        summarize = helpers.get_summarize_penilaian(request.user.id)
        context = {
            'title' : 'Penilaian Mitra',
            'jml_kegiatan' : summarize['jml_kegiatan'],
            'jml_penilaian_aktif' : summarize['jml_penilaian_aktif'],
            'jml_mitra_belum_dinilai' : summarize['jml_mitra_belum_dinilai'],
            'jml_mitra_dinilai' : summarize['jml_mitra_dinilai'],
        }

        return render(request, 'master_penilaian/entry_nilai.html', context)
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.PenilaianMitraForm(request.POST)
            if form.is_valid():
                df = form.cleaned_data
                objs_create, objs_update = 0, 0
                for dt in df:
                    modelPenilaian = models.MasterPenilaianPetugas
                    modelNilai = models.MasterNilaiPetugas
                    petugas = AlokasiPetugas.objects.get(pk = dt['petugas'])
                    penilai = AlokasiPetugas.objects.get(pk = dt['penilai'])
                    indikator_penilaian = models.IndikatorKegiatanPenilaian.objects.get(pk = dt['penilaian'])

                    role_permitted = list(indikator_penilaian.kegiatan_penilaian.role_permitted.values_list('id', flat=True))
                    role_permitted_penilai = list(indikator_penilaian.kegiatan_penilaian.role_penilai_permitted.values_list('id', flat=True))

                    if petugas.role.id not in role_permitted:
                        return JsonResponse({"status": "failed", "messages":  f'Data alokasi petugas dengan role {petugas.role.jabatan} tidak diizinkan untuk mengikuti kegiatan penilaian "<i>{indikator_penilaian.kegiatan_penilaian.nama_kegiatan}</i>"'})
                    
                    if penilai.role.id not in role_permitted_penilai:
                        return JsonResponse({"status": "failed", "messages":  f'Data alokasi pegawai dengan role {penilai.role.jabatan} tidak diizinkan untuk mengikuti kegiatan penilaian "<i>{indikator_penilaian.kegiatan_penilaian.nama_kegiatan}</i>"'})
            
                    nilai = dt['nilai']
                    catatan = dt['catatan']

                    db_check = modelPenilaian.objects.filter(petugas = petugas, penilai = penilai, detail_nilai__indikator_penilaian__kegiatan_penilaian = indikator_penilaian.kegiatan_penilaian).values('id', 'petugas', 'penilai').distinct()

                    if db_check.exists():
                        db_check2 = models.MasterNilaiPetugas.objects.filter(penilaian = db_check.first()['id'], indikator_penilaian = indikator_penilaian.pk)
                        if db_check2.exists():
                            nilai_mitra_update = db_check2.first()
                            nilai_mitra_update.nilai = nilai
                            nilai_mitra_update.catatan = catatan
                            nilai_mitra_update.save()
                            objs_update += 1
                        else:
                            modelNilai(
                                penilaian = modelPenilaian.objects.get(pk = db_check.first()['id']),
                                indikator_penilaian = indikator_penilaian,
                                nilai = nilai,
                                catatan = catatan,
                            ).save()
                            objs_create += 1
                    else:
                        row_affected = modelPenilaian.objects.create(petugas = petugas, penilai = penilai, state = '2')
                        modelNilai(
                            penilaian = row_affected,
                            indikator_penilaian = indikator_penilaian,
                            nilai = nilai,
                            catatan = catatan,
                        ).save()
                        objs_create += 1

                msg = ''
                if objs_create > 0:
                    msg += f"Data <strong>berhasil</strong> ditambahkan.<br>"
                if objs_update > 0:
                    msg += f"Data <strong>berhasil</strong> diperbarui.<br>"

                return JsonResponse({"status": "success", "messages":  msg})
            
            return JsonResponse({"error": form.errors}, status=400)
        else:
            return JsonResponse({"error": ""}, status=400)

class KegiatanPenilaianJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):
        data = self._datatables(request)
        return HttpResponse(json.dumps(data, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data_pegawai= model_pegawai.MasterPegawaiModel.objects.filter(user=request.user.id)
        if data_pegawai.exists():
            data = models.KegiatanPenilaianModel.objects.filter(status='1')

            if search:
                data = data.filter(
                    Q(kegiatan_survey__nama_kegiatan__icontains=search)|
                    Q(kegiatan_survey__survey__nama__icontains=search)|
                    Q(tgl_penilaian__icontains=search)|
                    Q(status__icontains=search)|
                    Q(role_permitted__jabatan__icontains=search)|
                    Q(role_penilai_permitted__jabatan__icontains=search)
                )
            data = data.exclude(Q(kegiatan_survey=None)|Q(tgl_penilaian=None)|Q(status=None)|Q(role_permitted=None)|Q(role_penilai_permitted=None))
            records_total = data.count()
            records_filtered = records_total

            if 'status_penilaian' not in order_col_name:
                data = data.order_by(order_col_name)

            paginator = Paginator(data, length)

            try:
                object_list = paginator.page(page_number).object_list
            except PageNotAnInteger:
                object_list = paginator.page(1).object_list
            except EmptyPage:
                object_list = paginator.page(1).object_list
            
            data = []
            for obj in object_list:
                aloc = model_petugas.AlokasiPetugas.objects.filter(pegawai = data_pegawai.first().pk, sub_kegiatan = obj.kegiatan_survey)
                if aloc.exists():
                    role_penilai_permitted = obj.role_penilai_permitted.values_list('id', flat=True)
                    if aloc.first().role.pk not in role_penilai_permitted:
                        continue

                    check_penilaian = models.MasterPenilaianPetugas.objects.filter(penilai = aloc.first().pk, detail_nilai__indikator_penilaian__kegiatan_penilaian = obj.pk)
                    if check_penilaian.exists():
                        state_penilaian = f'<span class="badge bg-success p-1"> Anda Sudah Menilai </span>'
                    else:
                        state_penilaian = f'<span class="badge bg-warning p-1"> Anda Belum Menilai </span>'

                    if obj.status == '0':
                        state = f'bg-danger'
                    elif obj.status == '1':
                        state = f'bg-primary'
                    else:
                        state = f'bg-success'

                    data.append(
                        {
                            'kegiatan_survey__nama_kegiatan': obj.kegiatan_survey.nama_kegiatan,
                            'kegiatan_survey__survey__nama': obj.kegiatan_survey.survey.nama,
                            'tgl_penilaian': obj.tgl_penilaian.strftime('%d %b %Y'),
                            'status' :  f'<span class="badge {state} p-1"> {obj.get_status_display()} </span>',
                            'status_penilaian' : state_penilaian,
                            'aksi': f'<a href="javascript:void(0);" class="action-icon" onclick="pushValuePenilaian({obj.id})"><i class="mdi mdi-square-edit-outline font-15"></i></a>'
                        }
                    )
                continue

            if len(data) > 0 and 'status_penilaian' in order_col_name:
                if '-' in order_col_name:
                    data = sorted(data, key=lambda x: x[order_col_name.replace("-", "")], reverse=True)
                else:
                    data = sorted(data, key=lambda x: x[order_col_name])
        else:
            records_total = 0
            records_filtered = 0
            data = []
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class MasterNilaiPetugasClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        data = self._datatables(request)
        return HttpResponse(json.dumps(data, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search_mitra')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        records_total = 0
        records_filtered = 0
        data = []
        
        data_pegawai= model_pegawai.MasterPegawaiModel.objects.filter(user=request.user.id)

        if data_pegawai.exists():
            dataKegiatan = models.KegiatanPenilaianModel.objects.filter(status='1')
            
            if datatables.get('kegiatan_penilaian'):
                dataKegiatan = dataKegiatan.filter(pk = datatables.get('kegiatan_penilaian'))

            dataset = []
            for dt_kegiatan in dataKegiatan:
                role_penilai_permitted = dt_kegiatan.role_penilai_permitted.values_list('id', flat=True)

                aloc = model_petugas.AlokasiPetugas.objects.filter(sub_kegiatan = dt_kegiatan.kegiatan_survey)
                aloc_pegawai = aloc.filter(pegawai = data_pegawai.first().pk)
                if aloc_pegawai.exists():
                    if aloc_pegawai.first().role.pk not in role_penilai_permitted:
                        continue
                        
                    aloc = aloc.filter(pegawai=None)
                    if datatables.get('search_mitra'):
                        aloc = aloc.filter(
                            Q(petugas__adm_id__region__icontains=search)|
                            Q(petugas__nama_petugas__icontains=search)|
                            Q(role__jabatan__icontains=search)
                        )

                    for dt in aloc:
                        role_permitted = list(dt_kegiatan.role_permitted.values_list('id', flat=True))
                        if dt.role.pk not in role_permitted:
                            continue

                        nilai_mitra = {
                            'id_alokasi' : dt.id,
                            'id_kegiatan_penilaian' : dt_kegiatan.id,
                            'kegiatan_penilaian' : dt_kegiatan.kegiatan_survey.nama_kegiatan,
                            'wilayah' : f'Desa {dt.petugas.adm_id.region}',
                            'nama' : dt.petugas.nama_petugas,
                            'role' : dt.role.jabatan,
                        }

                        # Salah
                        qry = models.MasterPenilaianPetugas.objects.filter(detail_nilai__indikator_penilaian__kegiatan_penilaian = dt_kegiatan.pk, penilai__pegawai = data_pegawai.first().pk, petugas = dt.pk)

                        if qry.exists():
                            qry = (qry
                                .values('detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey__nama_kegiatan', 'petugas__petugas__adm_id__region', 'petugas__petugas__nama_petugas', 'petugas__role__jabatan')
                                .order_by('petugas__petugas__nama_petugas')
                                .annotate(avg = Avg('detail_nilai__nilai'))
                                .first()
                            )
                            dataset.append({ **nilai_mitra, **{
                                'rerata' : f'{round(qry['avg'], 2)}',
                                'state' : '<span class="badge bg-success p-1"> Sudah Menilai </span>',
                                'aksi' : f'<a href="javascript:void(0);" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a>'
                            }})
                            continue
                        
                        dataset.append({ **nilai_mitra, **{
                            'rerata' : '0',
                            'state' : '<span class="badge bg-info p-1"> Belum Menilai </span>',
                            'aksi' : f'<a href="javascript:void(0);" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a>'
                        }})

                    records_total = len(dataset)
                    records_filtered = records_total

            paginator = Paginator(dataset, length)

            try:
                object_list = paginator.page(page_number).object_list
            except PageNotAnInteger:
                object_list = paginator.page(1).object_list
            except EmptyPage:
                object_list = paginator.page(1).object_list
            
            for obj in object_list:
                data.append(
                    {
                        'kegiatan_penilaian': obj['kegiatan_penilaian'],
                        'wilayah': obj['wilayah'],
                        'nama': obj['nama'],
                        'role' :  obj['role'],
                        'rerata' : obj['rerata'],
                        'state' : obj['state'],
                        'aksi': f'<a href="javascript:void(0);" class="action-icon" onclick="entryPenilaian({obj["id_kegiatan_penilaian"]}, {obj["id_alokasi"]})"><i class="mdi mdi-square-edit-outline font-15"></i></a>'
                    }
                )
        
        if '-' in order_col_name:
            data = sorted(data, key=lambda x: x[order_col_name.replace("-", "")], reverse=True)
        else:
            data = sorted(data, key=lambda x: x[order_col_name])

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class IndikatorPenilaianPetugasClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            id_alokasi = request.POST.get('id_alokasi')
            id_kegiatan = request.POST.get('id_kegiatan')
            
            data_petugas = model_petugas.AlokasiPetugas.objects.filter(id=id_alokasi)
            data_pegawai = model_pegawai.MasterPegawaiModel.objects.filter(user=request.user.id)
            dataKegiatan = models.KegiatanPenilaianModel.objects.filter(pk= id_kegiatan, status='1')
            
            if data_pegawai.exists() and dataKegiatan.exists() and data_petugas.exists():

                penilai = data_pegawai.first()
                data_petugas = data_petugas.first()
                dataKegiatan = dataKegiatan.first()
                role_penilai = model_petugas.AlokasiPetugas.objects.filter(sub_kegiatan=dataKegiatan.kegiatan_survey, pegawai=penilai.pk).first()

                input_selected = {
                    'role_penilai_opts' : f'<option value="{role_penilai.id}" selected>{role_penilai.role.jabatan}</option>',
                    'penilai_opts' : f'<option value="{role_penilai.id}" selected>[{penilai.nip}] {penilai.name}</option>',
                    'kegiatan_opts' : f'<option value="{dataKegiatan.kegiatan_survey.id}" selected>Penilaian {dataKegiatan.kegiatan_survey.nama_kegiatan}</option>',
                    'survei_opts' : f'<option value="{dataKegiatan.kegiatan_survey.survey.id}" selected>{dataKegiatan.kegiatan_survey.survey.nama}</option>',
                    'petugas_opts' : f'<option value="{data_petugas.pk}" selected>{data_petugas.petugas.nama_petugas}</option>',
                    'role_petugas_opts' : f'<option value="{data_petugas.role.id}" selected>{data_petugas.role.jabatan}</option>'
                }

                list_indicators = models.IndikatorKegiatanPenilaian.objects.filter(kegiatan_penilaian = dataKegiatan.pk)

                opt = ''
                for idx, dt in enumerate(list_indicators):
                    nilai_petugas = models.MasterPenilaianPetugas.objects.filter(penilai = role_penilai.pk, petugas = data_petugas.pk, detail_nilai__indikator_penilaian = dt.pk ).values('detail_nilai__nilai', 'detail_nilai__catatan')
                    nilai, catatan = '', ''
                    if nilai_petugas.exists():
                        nilai = nilai_petugas.first()['detail_nilai__nilai']
                        catatan = nilai_petugas.first()['detail_nilai__catatan']

                    opt +=  f"""<tr>
                                <td>{idx+1}</td>
                                <td>{dt.indikator_penilaian.nama_indikator}</td>
                                <td>
                                    <input type="number" class="form-control d-inline" name="nilai_indikator_{dt.pk}" value="{nilai}" placeholder="Isikan nilai" alt="Isikan nilai" min="{dt.n_min}" max="{dt.n_max}" onkeyup="if(this.value > {dt.n_max} || this.value < {dt.n_min}) this.value = null;">   
                                </td>
                                <td>
                                    <textarea class="form-control" name="catatan_indikator_{dt.pk}" cols="30" rows="5">{catatan}</textarea>
                                </td>
                            </tr>"""
                    
                input_selected['tbody'] = opt
                return JsonResponse({'status': 'success', 'data': input_selected}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class MasterGlobalRankPetugasClassView(LoginRequiredMixin, View):

    def post(self, request):
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = request.POST.get('search_mitra')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.MasterPenilaianPetugas.objects
        if search:
            data = data.filter(Q(petugas__petugas__adm_id__code__icontains = search) | Q(petugas__petugas__nama_petugas__icontains = search) | Q(petugas__petugas__kode_petugas__icontains = search) )

        data = (data.values('petugas__petugas__adm_id__code', 'petugas__petugas__nama_petugas')
            .annotate(avg = Avg('detail_nilai__nilai'))
            .order_by('-avg'))

        records_total = data.count()
        records_filtered = records_total
        
        nilai_petugas = []
        for idx, dt in enumerate(data):
            kec = AdministrativeModel.objects.filter(code = dt['petugas__petugas__adm_id__code'][:-3]).first()
            nilai_petugas.append({
                'wilayah' : kec.region,
                'petugas' : dt['petugas__petugas__nama_petugas'],
                'rerata' : round(dt['avg'], 2),
                'rank' : idx+1,
                'rank_' : f'{idx+1} / {records_total}',
            })
        
        if '-' in order_col_name:
            object_list = sorted(nilai_petugas, key=lambda x: x[order_col_name.replace("-", "")], reverse=True)
        else:
            object_list = sorted(nilai_petugas, key=lambda x: x[order_col_name])

        data = [
            {
                'wilayah': f'Kec. {obj["wilayah"]}',
                'petugas': obj['petugas'],
                'rerata': obj['rerata'],
                'rank': obj['rank_'],
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

