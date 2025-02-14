import json
from django.urls import reverse, reverse_lazy
from django.core.serializers.json import DjangoJSONEncoder

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.shortcuts import render
from django.views import View
from django.core import serializers
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from . import models, forms, utils

import itertools

from .resources import MasterPetugasResource, MasterAlokasiResource, MasterRoleResource
from master_survey.models import SurveyModel, SubKegiatanSurvei
from master_penilaian.models import MasterPenilaianPetugas, MasterNilaiPetugas, KegiatanPenilaianModel
from master_pegawai.models import MasterPegawaiModel

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from . import utils
from pprint import pprint

from statistics import mean
from operator import itemgetter
from django.shortcuts import redirect

from munapps.helpers import currency_formatting as cf, restrict_datatable_actions
from django.db.models.functions import Length

from master_honor.models import HonorModel
from master_pegawai.models import MasterPegawaiModel

from munapps.mixins import RestrictionsAccess, RestrictionsHttpRequestAccess
from munapps.helpers import get_adm_levels

class MasterPetugasJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)
        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.MasterPetugas.objects
        if datatables.get('adm_filter'):
            data = data.filter(Q(adm_id__code__icontains=datatables.get('adm_filter')))

        if datatables.get('jk_filter'):
            data = data.filter(Q(jk=datatables.get('jk_filter')))

        if datatables.get('agama_filter'):
            data = data.filter(Q(agama=datatables.get('agama_filter')))

        if datatables.get('bank_filter'):
            data = data.filter(Q(bank=datatables.get('bank_filter')))

        if datatables.get('pendidikan_filter'):
            data = data.filter(pendidikan = datatables.get('pendidikan_filter'))

        if datatables.get('status_filter'):
            data = data.filter(status = datatables.get('status_filter'))

        if search:
            data = data.filter(
                Q(adm_id__code=search)|
                Q(adm_id__region=search)|
                Q(kode_petugas__icontains=search)|
                Q(kode_petugas__icontains=search)|
                Q(nama_petugas__icontains=search)|
                Q(nik__icontains=search)|
                Q(email__icontains=search)|
                Q(no_telp__icontains=search)|
                Q(status__icontains=search)
            )

        records_total = data.count()
        records_filtered = records_total
        
        data = data.order_by(order_col_name)
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        restrict_actions = restrict_datatable_actions(request)
        data = []
        for obj in object_list:
            if obj.status == '0':
                class_badge = 'badge badge-success-lighten'
            elif obj.status == '1':
                class_badge = 'badge badge-secondary-lighten'
            elif obj.status == '3':
                class_badge = 'badge badge-danger-lighten'
            else:
                class_badge = 'badge badge-primary-lighten'
            kec = models.AdministrativeModel.objects.filter(code=obj.adm_id.code[:7]).first()
            adm = f'Kec. {kec.region}, {obj.adm_id.region}'
            data.append({
                        'adm_id__region': adm,
                        'kode_petugas': obj.kode_petugas,
                        'nama_petugas': f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": obj.id})}" class="text-body">{obj.nama_petugas}</a>',
                        'nik': obj.nik,
                        'email': obj.email,
                        'no_telp': obj.no_telp,
                        'status': f'<span class="badge {class_badge}"> {obj.get_status_display()} </span>',
                        'aksi': '-' if restrict_actions else f'<a href="javascript:void(0);" onclick="editPetugas({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="hapusPetugas({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>'
            })
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class MasterPetugasClassView(LoginRequiredMixin, View):

    def get(self, request):
        form = forms.MasterPetugasForm()
        context = {
            'title' : 'Master Mitra',
            'data_petugas' : models.MasterPetugas.objects.all(),
            'adm_prov' : models.AdministrativeModel.objects.annotate(text_len=Length('code')).filter(text_len=2).order_by('region'),
            'adm' : get_adm_levels(models.AdministrativeModel),
            'form': form,
            'form_upload': forms.MasterPetugasFormUpload()
            }
       
        return render(request, 'master_petugas/index.html', context)
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.MasterPetugasForm(request.POST)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)

class MasterPetugasDeleteView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')
                data_petugas = models.MasterPetugas.objects.filter(pk = id)
                if data_petugas.exists():
                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterPetugasDetailView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                data_petugas = models.MasterPetugas.objects.filter(pk=id)

                if data_petugas.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(data_petugas.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class MasterPetugasUpdateView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    data_petugas = models.MasterPetugas.objects

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            
            data = get_object_or_404(models.MasterPetugas, pk=request.POST.get('id'))

            form = forms.MasterPetugasForm(request.POST, instance=data)

            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


class MasterPetugasExportClassView(LoginRequiredMixin, View):
    def get(self, request):
    
        resource = MasterPetugasResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Master Petugas.xls"'
        return response 
    

class MasterPetugasTemplateClassView(LoginRequiredMixin, RestrictionsAccess, View):

    def get(self, request, *args, **kwargs): 

        def_rows = self.kwargs['rows']
        wb = Workbook()

        ws = wb.active

        # Ini untuk header columns
        ws.title = 'Form Upload Data Petugas'

        header = utils.get_verbose_fields(models.MasterPetugas, exclude_pk=True)
        header = ['No'] + header

        head_row = 2
        header_cols = np.array(ws[f'A{head_row}':f'R{head_row}'])

        # Set value and style for header
        for v,c in zip(header, header_cols.T.flatten()):
            # Set style
            c.font = Font(name='Cambria', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.value = v

        # Adjustment cols
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50

        for i in range(3, 500):
            for dt in ['B', 'C', 'F', 'G', 'H', 'M', 'Q'] :
                ws[f'{dt}{i}'].number_format = '@'

        ws.merge_cells('A1:R1')
        ws['A1'] = 'Template Upload Data Mitra'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Cambria',bold=True, size=14)

        for row in range(def_rows):
            ws[f'A{row+3}'] = row+1
            ws[f'A{row+3}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 2 for Metadata
        ws1 = wb.create_sheet('Metadata Formulir Pengisian')
        ws1.merge_cells('A1:F1')

        ws1['A1'] = 'Metadata'
        ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws1['A1'].font = Font(name='Cambria',bold=True, size=12)

        jk = list(models.MasterPetugas._meta.get_field('jk').choices)
        pendidikan = list(models.MasterPetugas._meta.get_field('pendidikan').choices)
        agama = list(models.MasterPetugas._meta.get_field('agama').choices)
        status = list(models.MasterPetugas._meta.get_field('status').choices)
        bank = list(models.MasterPetugas._meta.get_field('bank').choices)
        adm = models.AdministrativeModel.objects.annotate(text_len=Length('code')).filter(text_len=10).order_by('region')

        adm_lists = []
        for idx, dt in enumerate(adm):
            adm_lists.append((idx, f'[{dt.code}] {dt.region}'))

        utils.generate_meta_templates(ws1, 'A', 2, 'Jenis Kelamin', jk)
        utils.generate_meta_templates(ws1, 'B', 2, 'Data Pendidikan', pendidikan)
        utils.generate_meta_templates(ws1, 'C', 2, 'Agama', agama)
        utils.generate_meta_templates(ws1, 'D', 2, 'Status Mitra', status)
        utils.generate_meta_templates(ws1, 'E', 2, 'Jenis Bank', bank)
        utils.generate_meta_templates(ws1, 'F', 2, 'Daftar Desa', adm_lists)

        utils.generate_field_Validation(ws, ws1, 'A', 3, len(jk), 'E', 3, def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'B', 3, len(pendidikan), 'I', 3, def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'C', 3, len(agama), 'K',3,  def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'D', 3, len(status), 'O', 3,  def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'E', 3, len(bank), 'P', 3,  def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'F', 3, len(adm_lists), 'B', 3,  def_rows=def_rows)

        ws1.protection.password = "Bqlbz110"
        ws1.protection.sheet = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Data Mitra.xlsx'

        wb.save(response)
        return response

class MasterPetugasUploadClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
           
            form = forms.MasterPetugasFormUpload(request.POST, request.FILES)
            model = models.MasterPetugas
            admInstance = models.AdministrativeModel.objects

            if form.is_valid():
                df = form.cleaned_data
                objs = []
                for idx in range(len(df['id'])):
                    objs.append(
                        model(
                            adm_id = admInstance.filter(code = df['adm_id'][idx]).first(),
                            kode_petugas = df['kode_petugas'][idx],
                            nama_petugas= df['nama_petugas'][idx],
                            jk= df['jk'][idx],
                            nik= df['nik'][idx],
                            npwp= df['npwp'][idx],
                            tgl_lahir= df['tgl_lahir'][idx],
                            pendidikan= df['pendidikan'][idx],
                            pekerjaan= df['pekerjaan'][idx],
                            agama= df['agama'][idx],
                            email= df['email'][idx],
                            no_telp= df['no_telp'][idx],
                            alamat= df['alamat'][idx],
                            status = df['status'][idx],
                            bank = df['bank'][idx],
                            rekening = df['rekening'][idx],
                            pemilik_rek = df['pemilik_rek'][idx],
                        )
                    )

                model.objects.bulk_create(objs)
                return JsonResponse({"status": "success", "messages": f"<strong>{len(df['id'])}</strong> Data berhasil diupload."})
            else:
                
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})

        return JsonResponse({"error": ""}, status=403)
    
class MasterPetugasDetailViewClassView(LoginRequiredMixin, View):

    def _globalRank(self, request):
        data = MasterPenilaianPetugas.objects.values('petugas__petugas__kode_petugas', 'petugas__petugas__nama_petugas', 'petugas__sub_kegiatan__survey__nama', 'petugas__sub_kegiatan__survey__tgl_mulai', 'petugas__role__jabatan', 'detail_nilai__indikator_penilaian__kegiatan_penilaian','petugas__sub_kegiatan__nama_kegiatan', 'detail_nilai__nilai', 'detail_nilai__catatan')
        master_data = []
        for dt in data:
            check_exist = [index for (index, d) in enumerate(master_data) if d["kode_petugas"] == dt['petugas__petugas__kode_petugas']]
            
            if len(check_exist) > 0:
                check_exist_2 = [index for (index, d) in enumerate(master_data[check_exist[0]]['kegiatan_penilaian']) if d["id_kegiatan"] == dt['detail_nilai__indikator_penilaian__kegiatan_penilaian']]
                
                if len(check_exist_2) > 0:
                    master_data[check_exist[0]]['kegiatan_penilaian'][check_exist_2[0]]['nilai'].append(dt['detail_nilai__nilai'])
                    master_data[check_exist[0]]['kegiatan_penilaian'][check_exist_2[0]]['catatan'].append(dt['detail_nilai__catatan'])
                else:
                    master_data[check_exist[0]]['kegiatan_penilaian'].append({
                        'id_kegiatan' : dt['detail_nilai__indikator_penilaian__kegiatan_penilaian'],
                        'survey' : dt['petugas__sub_kegiatan__survey__nama'],
                        'tahun' : dt['petugas__sub_kegiatan__survey__tgl_mulai'],
                        'nama_kegiatan': dt['petugas__sub_kegiatan__nama_kegiatan'],
                        'role': dt['petugas__role__jabatan'],
                        'nilai' : [dt['detail_nilai__nilai']],
                        'catatan' : [dt['detail_nilai__catatan']],
                    })
                continue

            master_data.append({
                'kode_petugas': dt['petugas__petugas__kode_petugas'],
                'nama_petugas': dt['petugas__petugas__nama_petugas'],
                'rerata_final': 0,
                'ranking_final': 0,
                'kegiatan_penilaian' : [{'id_kegiatan': dt['detail_nilai__indikator_penilaian__kegiatan_penilaian'] , 'role' :  dt['petugas__role__jabatan'], 'survey' : dt['petugas__sub_kegiatan__survey__nama'], 'tahun' : dt['petugas__sub_kegiatan__survey__tgl_mulai'], 'nama_kegiatan': dt['petugas__sub_kegiatan__nama_kegiatan'], 'nilai': [dt['detail_nilai__nilai']], 'catatan': [dt['detail_nilai__catatan']]}]
            })

        for dt in master_data:
            mean_data = []
            for dt_kegiatan in dt['kegiatan_penilaian']:
                mean_data.append(round(mean(dt_kegiatan['nilai']), 2))
        
            dt['rerata_final'] = round(mean(mean_data), 2)

        data_sorted = sorted(master_data, key = itemgetter('rerata_final'), reverse=True)
        for idx, dt in enumerate(data_sorted):
            dt['ranking_final'] = idx+1

        return data_sorted

    def get(self, request, *args, **kwargs):
        
        mitra_id = self.kwargs['mitra_id']
        mitra = models.MasterPetugas.objects.filter(pk = mitra_id)

        if mitra.exists() == False:
            return redirect(request.META.get('HTTP_REFERER', '/'))

        survei_ = models.AlokasiPetugas.objects.filter(petugas = mitra_id)

        kegiatan_penilaian_ = MasterPenilaianPetugas.objects.filter(petugas__petugas = mitra_id).values('detail_nilai__indikator_penilaian__kegiatan_penilaian')

        global_rank = self._globalRank(request)

        # Mengurutkan penilaian berdasarkan kegiatan penilaian
        data_nilai_mitra = {}
        for dt in global_rank:
            dt_kode_petugas = dt['kode_petugas']
            dt_kegiatan_penilaian = dt['kegiatan_penilaian']

            for dt_ in dt_kegiatan_penilaian:

                dt_['rerata'] = round(mean(dt_['nilai']), 2)

                if dt_['id_kegiatan'] in data_nilai_mitra:
                    data_nilai_mitra[dt_['id_kegiatan']].append(dt_ | {'kode_petugas' : dt_kode_petugas})
                else:
                    data_nilai_mitra[dt_['id_kegiatan']] = [
                        dt_ | {'kode_petugas' : dt_kode_petugas}
                    ]

        for idx, val in data_nilai_mitra.items():
            sorted_ = sorted(val, key = itemgetter('rerata'), reverse=True)
            
            for idx2, dt_sort in enumerate(sorted_):
                dt_sort['rank'] = f'{idx2+1} of {len(sorted_)}'

            data_nilai_mitra[idx] = sorted_


        # Formatting Data
        data_final = []
        for dt_ in kegiatan_penilaian_.distinct():
            id_kegiatan_penilaian = dt_['detail_nilai__indikator_penilaian__kegiatan_penilaian']
            if id_kegiatan_penilaian in data_nilai_mitra:
                filter_data = [index for (index, d) in enumerate(data_nilai_mitra[id_kegiatan_penilaian]) if d["kode_petugas"] == mitra.first().kode_petugas]

                if len(filter_data) > 0:
                    data_final.append(data_nilai_mitra[id_kegiatan_penilaian][filter_data[0]])


        for dt in data_final:
            dt['catatan'] = np.unique(np.array(dt['catatan']))
        
        check_exist = [index for (index, d) in enumerate(global_rank) if d["kode_petugas"] == mitra.first().kode_petugas]
        global_ranking = global_rank[check_exist[0]]['ranking_final'] if len(check_exist) > 0 else ''

        context = {
            'title' : f'{mitra.first().kode_petugas} | {mitra.first().nama_petugas}',
            'mitra' : mitra.first(),
            'survei_followed' : survei_.count(),
            'history_survey' : survei_,
            'kegiatan_followed' : kegiatan_penilaian_.distinct().count(),
            'global_rank' : global_ranking,
            'penilaian' : data_final
        }

        return render(request, 'master_petugas/detail_petugas.html', context)

class MasterPetugasSearchClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:

            if request.method == 'POST':
                
                search_by = request.POST.get('search_by')
                data_petugas = models.MasterPetugas.objects.filter(
                    Q(kode_petugas__icontains=search_by)|
                    Q(nama_petugas__icontains=search_by)|
                    Q(nik__icontains=search_by)|
                    Q(email__icontains=search_by)|
                    Q(no_telp__icontains=search_by)
                ).exclude(Q(nama_petugas=None)|Q(nik=None)|Q(email=None)|Q(no_telp=None)|Q(status=None))

            row_counts = data_petugas.count()
            search_result = ''
            for dt in data_petugas[:5]:
                search_result += f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": dt.id})}" class="dropdown-item notify-item"><i class="mdi mdi-account-circle-outline me-1"></i><span>{dt.nama_petugas} ({dt.kode_petugas})</span></a>'
          
            return JsonResponse({'status' : 'success', 'search_result': {'row_count' : row_counts , 'search_result' : search_result }}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class MasterPetugasSearchClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:

            if request.method == 'POST':
                
                search_by = request.POST.get('search_by')
                data_petugas = models.MasterPetugas.objects.filter(
                    Q(kode_petugas__icontains=search_by)|
                    Q(nama_petugas__icontains=search_by)|
                    Q(nik__icontains=search_by)|
                    Q(email__icontains=search_by)|
                    Q(no_telp__icontains=search_by)
                ).exclude(Q(nama_petugas=None)|Q(nik=None)|Q(email=None)|Q(no_telp=None)|Q(status=None))

            row_counts = data_petugas.count()
            search_result = ''
            for dt in data_petugas[:5]:
                search_result += f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": dt.id})}" class="dropdown-item notify-item"><i class="mdi mdi-account-circle-outline me-1"></i><span>{dt.nama_petugas} ({dt.kode_petugas})</span></a>'
          
            return JsonResponse({'status' : 'success', 'search_result': {'row_count' : row_counts , 'search_result' : search_result }}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class GetAdministratifLocClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                next_code_level = 2
                code = request.POST.get('code')
                if len(code) == 2:
                    opts = '<option value="">-- Pilih Kab/Kota --</option>'
                    next_code_level = 4 
                elif len(code) == 4:
                    opts = '<option value="">-- Pilih Kecamatan --</option>'
                    next_code_level = 7 
                else:
                    opts = '<option value="">-- Pilih Desa --</option>'
                    next_code_level = 10 

                adm = models.AdministrativeModel.objects.annotate(text_len=Length('code')).filter(Q(code__icontains=code)).filter(text_len=next_code_level).order_by('code')
                for dt in adm:
                    opts += f'<option value="{dt.code}">[{dt.code}] {dt.region}</option>'
        
                return JsonResponse({'status' : 'success', 'adm' : opts}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


### Alokasi Petugas

class AlokasiPenugasanClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    
    def get(self, request):
        context = {
            'title' : 'Alokasi Penugasan',
            'form' : forms.AlokasiForm()
            }

        return render(request, 'master_petugas/alokasi_penugasan.html', context)
    
class AlokasiPetugasClassView(LoginRequiredMixin, RestrictionsAccess, View):
    
    def get(self, request):
        context = {
            'title' : 'Alokasi Mitra',
            'data_mitra': models.MasterPetugas.objects.filter(~Q(status = 3), ~Q(status = 1)),
            'data_pegawai': MasterPegawaiModel.objects.all(),
            'data_survei' : SurveyModel.objects.all(),
            'data_sub_kegiatan' : SubKegiatanSurvei.objects.filter(status = 1),
            'data_jabatan' : models.RoleMitra.objects.all(),
            'data_batasan_honor' : HonorModel.objects.all(),
            'form_upload' : forms.AlokasiPetugasFormUpload(),
            'form' : forms.AlokasiForm()
            }

        return render(request, 'master_petugas/alokasi_peran.html', context)


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax and request.method == 'POST':
            form = forms.AlokasiForm(request.POST)
            if form.is_valid():
                if request.POST.get('petugas') or request.POST.get('pegawai'):
                    if request.POST.get('petugas') :
                        data = get_object_or_404(models.MasterPetugas, pk=form.data['petugas'])
                        if data.status == '1' or data.status == '3':
                            return JsonResponse({"status":"failed", 'message': f'Mitra dengan status {data.get_status_display()} tidak dapat dialokasikan.'}, status=200)

                        check_exist_data = models.AlokasiPetugas.objects.filter(petugas = form.data['petugas'], sub_kegiatan = form.data['sub_kegiatan'])
                    else:
                        data = get_object_or_404(models.MasterPegawaiModel, pk=form.data['pegawai'])
                        check_exist_data = models.AlokasiPetugas.objects.filter(pegawai = form.data['pegawai'], sub_kegiatan = form.data['sub_kegiatan'])
                        
                    if check_exist_data.exists():
                        return JsonResponse({"status":"failed", 'message': 'Data alokasi telah tersedia pada database'}, status=200)
                    
                    form.save()

                    return JsonResponse({"status":"success", 'message': 'Data berhasil ditambahkan'}, status=200)
                else:
                    return JsonResponse({"status":"failed", 'message': 'Data alokasi gagal ditambahkan'}, status=200)

            else:
                return JsonResponse({"status":"failed", "error": form.errors}, status=400)
        return JsonResponse({"status":"failed", "error": ""}, status=400)
    

class AlokasiPetugasDeleteView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')
                data_petugas = models.AlokasiPetugas.objects.filter(pk = id)
                if data_petugas.exists():
                    check_nilai_mitra = MasterPenilaianPetugas.objects.filter(Q(petugas = id) | Q(penilai = id))
                    if check_nilai_mitra.exists() is False:
                        data_petugas.delete()
                        return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)

                    return JsonResponse({'status' : 'failed', 'message': 'Data penugasan telah digunakan pada master data penilaian.'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data petugas tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterAlokasiDetailView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                data_petugas = models.AlokasiPetugas.objects.filter(pk=id)

                # check_nilai_mitra = MasterNilaiPetugas.objects.filter(petugas = id)
                # if check_nilai_mitra.exists():
                #     return JsonResponse({'status' : 'failed', 'message': 'Data alokasi petugas telah digunakan pada master data penilaian'}, status=200)
                
                if data_petugas.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(data_petugas.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 


class MasterAlokasiUpdateView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax and request.method == 'POST':
            
            if request.POST.get('petugas') or request.POST.get('pegawai'):
                
                data = get_object_or_404(models.AlokasiPetugas, pk=request.POST.get('id'))
                safe_check = MasterPenilaianPetugas.objects.filter(Q(petugas = request.POST.get('id')) | Q(penilai = request.POST.get('id')))
                if safe_check.exists():
                    return JsonResponse({'status' : 'failed', 'message': 'Data penugasan yang telah terdaftar pada master data penilaian tidak dapat diupdate.'}, status=200)

                form = forms.AlokasiForm(request.POST, instance=data)

                if form.is_valid():
                    if request.POST.get('petugas') :
                        if data.petugas.status == '1' or data.petugas.status == '3':
                            return JsonResponse({"status":"failed", 'message': f'Mitra dengan status {data.petugas.get_status_display()} tidak dapat dialokasikan.'}, status=200)
                        check_exist_data = models.AlokasiPetugas.objects.filter(~Q(pk=form.data['id']) & Q(petugas = form.data['petugas']) & Q(sub_kegiatan = form.data['sub_kegiatan']))
                    else:
                        check_exist_data = models.AlokasiPetugas.objects.filter(~Q(pk=form.data['id']) & Q(pegawai = form.data['pegawai']) & Q(sub_kegiatan = form.data['sub_kegiatan']))

                    if check_exist_data.exists():
                        return JsonResponse({"status":"failed", 'message': 'Data penugasan telah tersedia pada database'}, status=200)

                    instance = form.save()
                    ser_instance = serializers.serialize('json', [ instance, ])
                    
                    return JsonResponse({"status":"success", "instance": ser_instance, "message": "Data berhasil diubah"}, status=200)
                else:
                    return JsonResponse({"status": "failed", "error": form.errors, "message": "Terjadi kesalahan"}, status=400)
            else:
                return JsonResponse({"status":"failed", 'message': f'Tidak dapat mengupdate data alokasi mitra.'}, status=200)
        
            
        return JsonResponse({"status": "failed", "message": "Terjadi Kesalahan"}, status=400)


class MasterAlokasiJsonResponseClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.AlokasiPetugas.objects
        if datatables.get('survei_filter'):
            data = data.filter(sub_kegiatan__survey = datatables.get('survei_filter'))

        if datatables.get('jabatan_filter'):
            data = data.filter(role = datatables.get('jabatan_filter'))
        
        if search:
            data = models.AlokasiPetugas.objects.filter(
                Q(pegawai__name__icontains=search)|
                Q(pegawai__nip__icontains=search)|
                Q(pegawai__nip_bps__icontains=search)|
                Q(petugas__kode_petugas__icontains=search)|
                Q(petugas__nama_petugas__icontains=search)|
                Q(sub_kegiatan__survey__nama__icontains=search)|
                Q(sub_kegiatan__nama_kegiatan__icontains=search)|
                Q(role__jabatan__icontains=search)
            )

        data = data.exclude((Q(petugas=None) & Q(pegawai=None))|Q(role=None)|Q(sub_kegiatan=None))
        records_total = data.count()
        records_filtered = records_total
        
        data = data.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = []
        for obj in object_list:
            code = obj.pegawai.nip if obj.pegawai else obj.petugas.kode_petugas 
            name = obj.pegawai.name if obj.pegawai else obj.petugas.nama_petugas 
            state_mitra = 0 if obj.pegawai else 1
            
            data.append({
                'petugas__kode_petugas': code ,
                'petugas__nama_petugas': f'<a href="javascript:void(0)" class="text-body">{name}</a>' if obj.pegawai else f'<a href="{reverse_lazy("master_petugas:mitra-view-detail", kwargs={"mitra_id": obj.id})}" class="text-body" target="_blank">{name}</a>',
                'sub_kegiatan__nama_kegiatan': obj.sub_kegiatan.nama_kegiatan,
                'role__jabatan': obj.role.jabatan,
                'pegawai': '<span class="badge bg-primary p-1">Organik</span>' if obj.pegawai else '<span class="badge bg-info p-1">Mitra</span>',
                'aksi': f'<a href="javascript:void(0);" onclick="editAlokPetugas({obj.id}, {state_mitra})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="deleteAlokasi({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>'
            })

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


class MasterAlokasiExportClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    def get(self, request):
    
        resource = MasterAlokasiResource()
        dataset = resource.export()

        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Alokasi Petugas.xls"'
        return response 

class MasterAlokasiTemplateClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def get(self, request):
        if request.GET.get('rows') is None:
            return redirect(reverse('master_petugas:alokasi'))
        
        def_rows = int(request.GET.get('rows'))
        wb = Workbook()
        ws = wb.active

        # Ini untuk header columns
        ws.title = 'Upload Alokasi Petugas'

        header = utils.get_verbose_fields(models.AlokasiPetugas, exclude_pk=True)
        header = ['No'] + header

        head_row = 2
        header_cols = np.array(ws[f'A{head_row}':f'M{head_row}'])

        # Set value and style for header
        for v,c in zip(header, header_cols.T.flatten()):
            c.font = Font(name='Cambria', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.value = v

        # Adjustment cols
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50

        ws.merge_cells('A1:D1')
        ws['A1'] = 'Template Upload Alokasi Mitra'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Cambria',bold=True, size=14)

        for row in range(def_rows):
            ws[f'A{row+3}'] = row+1
            ws[f'A{row+3}'].alignment = Alignment(horizontal='center', vertical='center')

        # Sheet 2 for Metadata
        ws1 = wb.create_sheet('Metadata Formulir Pengisian')
        ws1.merge_cells('A1:D1')
        ws1['A1'] = 'Metadata'
        ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws1['A1'].font = Font(name='Cambria',bold=True, size=12)

        pegawai_lists = list(MasterPegawaiModel.objects.values_list('nip', 'name'))
        pegawai_choices = []
        for dt in pegawai_lists:
            pegawai_choices.append((dt[0], f'[{dt[0]}] {dt[1]}'))
            
        mitra_lists = list(models.MasterPetugas.objects.filter(~Q(status=1), ~Q(status=3)).values_list('kode_petugas', 'nama_petugas'))
        mitra_choices = []
        for dt in mitra_lists:
            mitra_choices.append((dt[0], f'[{dt[0]}] {dt[1]}'))

        survey_lists = list(SubKegiatanSurvei.objects.filter(status='1').values_list('id','nama_kegiatan'))
        survey_choices = []
        for dt in survey_lists:
            survey_choices.append((dt[0], dt[1]))
        
        role_lists = list(models.RoleMitra.objects.values_list('id','jabatan'))
        role_choices = []
        for dt in role_lists:
            role_choices.append((dt[0], dt[1]))

        utils.generate_meta_templates(ws1, 'A', 2, 'Data Pegawai', pegawai_choices)
        utils.generate_meta_templates(ws1, 'B', 2, 'Data Mitra', mitra_choices)
        utils.generate_meta_templates(ws1, 'C', 2, 'Jabatan Mitra', role_choices)
        utils.generate_meta_templates(ws1, 'D', 2, 'Data Kegiatan Survei/Sensus', survey_choices)

        utils.generate_field_Validation(ws, ws1, 'A', 3, len(pegawai_choices), 'B', 3, def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'B', 3, len(mitra_choices), 'C', 3, def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'C', 3, len(role_choices), 'D', 3, def_rows=def_rows)
        utils.generate_field_Validation(ws, ws1, 'D', 3, len(survey_choices), 'E', 3, def_rows=def_rows)

        ws1.protection.password = "Bqlbz110"
        ws1.protection.sheet = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Alokasi Mitra.xlsx'

        wb.save(response)
        return response


class MasterAlokasiUploadClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            form = forms.AlokasiPetugasFormUpload(request.POST, request.FILES)
            model = models.AlokasiPetugas
            model_pegawai = MasterPegawaiModel
            model_mitra = models.MasterPetugas
            model_kegiatan = SubKegiatanSurvei
            model_role = models.RoleMitra

            if form.is_valid():
                df = form.cleaned_data
                objs = []
                for idx in range(len(df['id'])):
                    if np.isnan(df['pegawai'][idx]):
                        objs.append(
                            model(
                                petugas = model_mitra.objects.get(pk = df['petugas'][idx]),
                                sub_kegiatan=  model_kegiatan.objects.get(pk = df['sub_kegiatan'][idx]),
                                role = model_role.objects.get(pk = df['role'][idx])
                            )
                        )
                    else:
                        objs.append(
                            model(
                                pegawai = model_pegawai.objects.get(pk = df['pegawai'][idx]),
                                sub_kegiatan=  model_kegiatan.objects.get(pk = df['sub_kegiatan'][idx]),
                                role = model_role.objects.get(pk = df['role'][idx])
                            )
                        )
                model.objects.bulk_create(objs)
                return JsonResponse({"status": "success", "messages": f"<strong></strong> Data berhasil diupload."})
            else:
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})

        return JsonResponse({"error": ""}, status=403)  


# Role Petugas

class RolePetugasClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    
    def get(self, request):
        context = {
            'title' : 'Role Petugas',
            'form' : forms.RoleForm()
        }

        return render(request, 'master_petugas/role.html', context)


    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.RoleForm(request.POST)
            if form.is_valid():
                instance = form.save()
                user_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": user_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    

class MasterRoleJsonResponseClassView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for 
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.RoleMitra.objects
        if search:
            data = models.RoleMitra.objects.filter(
                Q(jabatan__icontains=search)
            )

        records_total = data.count()
        records_filtered = records_total
        
        data = data.order_by(order_col_name)
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'jabatan': obj.jabatan,
                'aksi':f'<a href="javascript:void(0);" onclick="editRole({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="deleteRole({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>',
            } for obj in object_list
        ]
        
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class RolePetugasDeleteView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')
                
                data_petugas = models.RoleMitra.objects.filter(pk = id)
                if data_petugas.exists():
                    check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(role_permitted = id)
                    if check_kegiatan_penilaian.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)
                
                    check_nilai_mitra = MasterPenilaianPetugas.objects.filter(Q(petugas__role = id) | Q(penilai__role = id))
                    if check_nilai_mitra.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)

                    data_petugas.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)

class MasterRoleDetailView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')
                role = models.RoleMitra.objects.filter(pk=id)

                if role.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(role.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class MasterRoleUpdateView(LoginRequiredMixin, RestrictionsHttpRequestAccess, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            data = get_object_or_404(models.RoleMitra, pk=request.POST.get('id'))
            check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(role_permitted = request.POST.get('id'))
            if check_kegiatan_penilaian.exists():
                return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)
            
            check_nilai_mitra = MasterPenilaianPetugas.objects.filter(petugas__role = request.POST.get('id'))
            if check_nilai_mitra.exists():
                return JsonResponse({'status' : 'failed', 'message': 'Data role petugas telah digunakan pada master data penilaian'}, status=200)

            form = forms.RoleForm(request.POST, instance=data)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({"status" : "success", "instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"status" : "failed", "error": form.errors, "message": "Terjadi Kesalahan"}, status=400)
        return JsonResponse({"status" : "failed", "message": "Terjadi Kesalahan"}, status=400)

class MasterRoleExportClassView(LoginRequiredMixin, RestrictionsAccess, View):
    def get(self, request):
        resource = MasterRoleResource()
        dataset = resource.export()
        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Jabatan Petugas.xls"'
        return response

class PetugasClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        master_mitra = models.MasterPetugas.objects.all()
        dataset = []
        for dt in master_mitra:
            alokasi_petugas = models.AlokasiPetugas.objects.filter(petugas = dt.pk)
            jml_kegiatan = 0
            jml_penilai = 0
            if alokasi_petugas.exists():
                jml_kegiatan = alokasi_petugas.count()
                for dt_ in alokasi_petugas:
                    check_penilaian = MasterPenilaianPetugas.objects.filter(petugas = dt_.id, detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey = dt_.sub_kegiatan).values('penilai_id').distinct()
                    if check_penilaian.exists():
                        jml_penilai += check_penilaian.count()

            dataset.append(
                {
                    'kode_petugas' : dt.kode_petugas,
                    'nama_petugas' : dt.nama_petugas,
                    'adm_id' : dt.adm_id.region,
                    'no_telp' : dt.no_telp,
                    'email' : dt.email,
                    'jml_kegiatan' : jml_kegiatan,
                    'jml_penilai' : jml_penilai,
                    'status' : '<span class="badge bg-success p-1"> Mitra Rekomendasi </span>'
                }
            )

        context = {
            'title' : 'Data Mitra',
        }

        return render(request, 'master_petugas/petugas.html', context)

class DetailPetugasPreviewClassView(LoginRequiredMixin, View):
    
    def _globalRank(self, request):
        data = MasterPenilaianPetugas.objects.values('petugas__petugas__kode_petugas', 'petugas__petugas__nama_petugas', 'petugas__sub_kegiatan__survey__nama', 'petugas__sub_kegiatan__survey__tgl_mulai', 'petugas__role__jabatan', 'detail_nilai__indikator_penilaian__kegiatan_penilaian','petugas__sub_kegiatan__nama_kegiatan', 'detail_nilai__nilai', 'detail_nilai__catatan')
        master_data = []
        for dt in data:
            check_exist = [index for (index, d) in enumerate(master_data) if d["kode_petugas"] == dt['petugas__petugas__kode_petugas']]
            
            if len(check_exist) > 0:
                check_exist_2 = [index for (index, d) in enumerate(master_data[check_exist[0]]['kegiatan_penilaian']) if d["id_kegiatan"] == dt['detail_nilai__indikator_penilaian__kegiatan_penilaian']]
                
                if len(check_exist_2) > 0:
                    master_data[check_exist[0]]['kegiatan_penilaian'][check_exist_2[0]]['nilai'].append(dt['detail_nilai__nilai'])
                    master_data[check_exist[0]]['kegiatan_penilaian'][check_exist_2[0]]['catatan'].append(dt['detail_nilai__catatan'])
                else:
                    master_data[check_exist[0]]['kegiatan_penilaian'].append({
                        'id_kegiatan' : dt['detail_nilai__indikator_penilaian__kegiatan_penilaian'],
                        'survey' : dt['petugas__sub_kegiatan__survey__nama'],
                        'tahun' : dt['petugas__sub_kegiatan__survey__tgl_mulai'],
                        'nama_kegiatan': dt['petugas__sub_kegiatan__nama_kegiatan'],
                        'role': dt['petugas__role__jabatan'],
                        'nilai' : [dt['detail_nilai__nilai']],
                        'catatan' : [dt['detail_nilai__catatan']],
                    })
                continue

            master_data.append({
                'kode_petugas': dt['petugas__petugas__kode_petugas'],
                'nama_petugas': dt['petugas__petugas__nama_petugas'],
                'rerata_final': 0,
                'ranking_final': 0,
                'kegiatan_penilaian' : [{'id_kegiatan': dt['detail_nilai__indikator_penilaian__kegiatan_penilaian'] , 'role' :  dt['petugas__role__jabatan'], 'survey' : dt['petugas__sub_kegiatan__survey__nama'], 'tahun' : dt['petugas__sub_kegiatan__survey__tgl_mulai'], 'nama_kegiatan': dt['petugas__sub_kegiatan__nama_kegiatan'], 'nilai': [dt['detail_nilai__nilai']], 'catatan': [dt['detail_nilai__catatan']]}]
            })

        for dt in master_data:
            mean_data = []
            for dt_kegiatan in dt['kegiatan_penilaian']:
                mean_data.append(round(mean(dt_kegiatan['nilai']), 2))
        
            dt['rerata_final'] = round(mean(mean_data), 2)

        data_sorted = sorted(master_data, key = itemgetter('rerata_final'), reverse=True)
        for idx, dt in enumerate(data_sorted):
            dt['ranking_final'] = idx+1

        return data_sorted


    def get(self, request, *args, **kwargs):
        mitra_id = self.kwargs['mitra_id']
        mitra = models.MasterPetugas.objects.filter(pk = mitra_id)

        if mitra.exists() == False:
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        survei_ = models.AlokasiPetugas.objects.filter(petugas = mitra_id)
        global_rank = self._globalRank(request)

        # Mengurutkan penilaian berdasarkan kegiatan penilaian
        data_nilai_mitra = {}
        for dt in global_rank:
            dt_kode_petugas = dt['kode_petugas']
            dt_kegiatan_penilaian = dt['kegiatan_penilaian']

            for dt_ in dt_kegiatan_penilaian:
                dt_['rerata'] = round(mean(dt_['nilai']), 2)
                if dt_['id_kegiatan'] in data_nilai_mitra:
                    data_nilai_mitra[dt_['id_kegiatan']].append(dt_ | {'kode_petugas' : dt_kode_petugas})
                else:
                    data_nilai_mitra[dt_['id_kegiatan']] = [
                        dt_ | {'kode_petugas' : dt_kode_petugas}
                    ]

        for idx, val in data_nilai_mitra.items():
            sorted_ = sorted(val, key = itemgetter('rerata'), reverse=True)
            
            for idx2, dt_sort in enumerate(sorted_):
                dt_sort['rank'] = f'{idx2+1} of {len(sorted_)}'

            data_nilai_mitra[idx] = sorted_

        # Formatting Data
        kegiatan_penilaian_ = MasterPenilaianPetugas.objects.filter(petugas__petugas = mitra_id).values('detail_nilai__indikator_penilaian__kegiatan_penilaian')
        data_final = []
        
        for dt_ in kegiatan_penilaian_.distinct():
            id_kegiatan_penilaian = dt_['detail_nilai__indikator_penilaian__kegiatan_penilaian']
            if id_kegiatan_penilaian in data_nilai_mitra:
                filter_data = [index for (index, d) in enumerate(data_nilai_mitra[id_kegiatan_penilaian]) if d["kode_petugas"] == mitra.first().kode_petugas]
                if len(filter_data) > 0:
                    data_final.append(data_nilai_mitra[id_kegiatan_penilaian][filter_data[0]])

        for dt in data_final:
            dt['catatan'] = np.unique(np.array(dt['catatan']))
        
        check_exist = [index for (index, d) in enumerate(global_rank) if d["kode_petugas"] == mitra.first().kode_petugas]
        global_ranking = global_rank[check_exist[0]]['ranking_final'] if len(check_exist) > 0 else ''

        context = {
            'title' : f'{mitra.first().kode_petugas} | {mitra.first().nama_petugas}',
            'mitra' : mitra.first(),
            'survei_followed' : survei_.count(),
            'kegiatan_followed' : kegiatan_penilaian_.distinct().count(),
            'global_rank' : global_ranking,
            'penilaian' : data_final
        }

        return render(request, 'master_petugas/detail_petugas_preview.html', context)

class ListPetugasClassView(LoginRequiredMixin, View):
    
    def post(self, request):    
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)
        search = request.POST.get('search_mitra')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for 
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data = models.MasterPetugas.objects.all()
        if search:
            data = data.filter(
                Q(adm_id__code__icontains=search)|
                Q(adm_id__region__icontains=search)|
                Q(kode_petugas__icontains=search)|
                Q(nama_petugas__icontains=search)|
                Q(nik__icontains=search)|
                Q(email__icontains=search)|
                Q(no_telp__icontains=search)
            )

        dataset = []
        for dt in data:
            alokasi_petugas = models.AlokasiPetugas.objects.filter(petugas = dt.pk)
            state_petugas = 'bg-success'
            if dt.status == '1':
                state_petugas = 'bg-warning'
            elif dt.status == '2':
                state_petugas = 'bg-primary'
            elif dt.status == '3':
                state_petugas = 'bg-danger'

            jml_kegiatan = 0
            jml_penilai = 0
            if alokasi_petugas.exists():
                jml_kegiatan = alokasi_petugas.count()

                for dt_ in alokasi_petugas:
                    check_penilaian = MasterPenilaianPetugas.objects.filter(petugas = dt_.id, detail_nilai__indikator_penilaian__kegiatan_penilaian__kegiatan_survey = dt_.sub_kegiatan).values('penilai_id').distinct()
                    if check_penilaian.exists():
                        jml_penilai += check_penilaian.count()

            dataset.append(
                {
                    'id' : dt.id,
                    'kode_petugas' : dt.kode_petugas,
                    'nama_petugas' : dt.nama_petugas,
                    'adm_id__region' : dt.adm_id.region,
                    'no_telp' : dt.no_telp if dt.no_telp else '-',
                    'email' : dt.email,
                    'jml_kegiatan' : jml_kegiatan,
                    'jml_penilai' : jml_penilai,
                    'status' : f'<span class="badge {state_petugas} p-1"> {dt.get_status_display()} </span>'
                }
            )

        records_total = len(dataset)
        records_filtered = records_total

        if '-' in order_col_name:
            data = sorted(dataset, key=lambda x: x[order_col_name.replace("-", "")], reverse=True)
        else:
            data = sorted(dataset, key=lambda x: x[order_col_name])
        
        # Conf Paginator
        paginator = Paginator(data, length)
        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = [
            {
                'kode_petugas' : obj['kode_petugas'],
                'nama_petugas' : obj['nama_petugas'],
                'adm_id__region' : obj['adm_id__region'],
                'no_telp' : obj['no_telp'],
                'email' : obj['email'],
                'jml_kegiatan' : obj['jml_kegiatan'],
                'jml_penilai' : obj['jml_penilai'],
                'status' : obj['status'],
                'aksi': f'<a href="{reverse_lazy("master_petugas:detail-petugas", kwargs={"mitra_id": obj["id"]})}" target="_blank" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a>',
            } for obj in object_list
        ]
        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }


