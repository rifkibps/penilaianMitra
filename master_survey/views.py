import json
from django.core.serializers.json import DjangoJSONEncoder

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.views import View
from django.core import serializers
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse

from .resources import MasterSurveiResource
from . import models
from . import forms

import itertools
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from master_petugas import utils
from master_petugas.models import AlokasiPetugas
from master_penilaian.models import MasterNilaiPetugas, KegiatanPenilaianModel, MasterPenilaianPetugas
from pprint import pprint

from munapps.helpers import currency_formatting as cf
# Create your views here.
class SurveyJsonResponseClassView(LoginRequiredMixin, View):

    def post(self, request):
        
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        
        # Get Draw
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)

        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]')) # Default 1st index for
        order_dir = datatables.get('order[0][dir]') # Descending or Ascending
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data_survey = models.SurveyModel.objects
        if datatables.get('state_filter'):
            data_survey = data_survey.filter(state = datatables.get('state_filter'))

        if search:
            data_survey = data_survey.filter(
                Q(nama__icontains=search)|Q(deskripsi__icontains=search)|Q(tgl_mulai__icontains=search)|Q(tgl_selesai__icontains=search)|Q(salary=search)
            )
        
        data_survey = data_survey.exclude(Q(nama=None)|Q(deskripsi=None)|Q(tgl_mulai=None)|Q(tgl_selesai=None))
        records_total = data_survey.count()
        records_filtered = records_total
        
        data_survey = data_survey.order_by(order_col_name)
        # Conf Paginator
        paginator = Paginator(data_survey, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        
        data = []

        for obj in object_list:
            if obj.state == '0':
                bg_class = 'bg-danger'
            elif obj.state == '1':
                bg_class = 'bg-info'
            else:
                bg_class = 'bg-success'

            data.append(
            {
                'nama': obj.nama,
                'tgl_mulai': obj.tgl_mulai.strftime('%d %B %Y') + ' s.d. ' + obj.tgl_selesai.strftime('%d %B %Y'),
                'deskripsi': obj.deskripsi,
                'state': f'<span class="badge {bg_class} p-1">{obj.get_state_display()}</span>',
                'aksi': f'<a href="javascript:void(0);" onclick="editSurvei({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="deleteSurvei({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>'
            })

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

class MasterSurveiClassView(LoginRequiredMixin, View):

    def get(self, request):
        form = forms.SurveiForm()
        context = {
            'title' : 'Master Survei',
            'form': form,
            'form_upload': forms.SurveiFormUpload(),
        }
        return render(request, 'master_survey/index.html', context)

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            form = forms.SurveiForm(request.POST)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    
class MasterSurveyUpdateView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:

            data = get_object_or_404(models.SurveyModel, pk=request.POST.get('id'))

            form = forms.SurveiForm(request.POST, instance=data)

            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                
                # send to client side.
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


class MasterSurveyDetailView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                
                id = request.POST.get('id')
                data_petugas = models.SurveyModel.objects.filter(pk=id)
                if data_petugas.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(data_petugas.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class MasterSurveyDeleteView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                pk = request.POST.get('id')
                data_survei = models.SurveyModel.objects.filter(pk = pk)
                
                if data_survei.exists():
                    check_alokasi_mitra = AlokasiPetugas.objects.filter(sub_kegiatan__survey = pk)
                    if check_alokasi_mitra.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada alokasi petugas.'}, status=200)
                    
                    check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(kegiatan_survey__survey = pk)
                    if check_kegiatan_penilaian.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada master data penilaian.'}, status=200)
                
                    data_survei.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterSurveyExportView(LoginRequiredMixin, View):

    def get(self, request):
        resource = MasterSurveiResource()
        dataset = resource.export()
        response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Master Survei.xls"'
        return response 
    

class MasterSurveyTemplateClassView(LoginRequiredMixin, View):

    def get(self, request,  *args, **kwargs): 
        if request.GET.get('rows') is None:
            return redirect(reverse('master_survei:index'))
        
        def_rows = int(request.GET.get('rows'))
        wb = Workbook()
        ws = wb.active
        # Ini untuk header columns
        ws.title = 'Upload Data Survei'

        header = utils.get_verbose_fields(models.SurveyModel, exclude_pk=True)
        header = ['No'] + header

        head_row = 2
        header_cols = np.array(ws[f'A{head_row}':f'M{head_row}'])

        # Set value and style for header
        for v,c in zip(header, header_cols.T.flatten()):
            # Set style
            c.font = Font(name='Cambria', size=12)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.value = v

        # Adjustment cols
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 22.50

        ws.merge_cells('A1:F1')
        ws['A1'] = 'Template Upload Kegiatan Pendataan'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Cambria',bold=True, size=14)
        
        for row in range(def_rows):
            ws[f'A{row+3}'] = row+1
            ws[f'A{row+3}'].alignment = Alignment(horizontal='center', vertical='center')

        # Sheet 2 for Metadata
        ws1 = wb.create_sheet('Metadata Formulir Pengisian')

        ws1['A1'] = 'Metadata'
        ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws1['A1'].font = Font(name='Cambria',bold=True, size=12)

        state_choices = list(models.SurveyModel._meta.get_field('state').choices)

        utils.generate_meta_templates(ws1, 'A', 2, 'Status Survei', state_choices)
        utils.generate_field_Validation(ws, ws1, 'A', 3, len(state_choices), 'F', 3, def_rows=def_rows)
        ws1.protection.password = "Bqlbz110"
        ws1.protection.sheet = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Template Upload Kegiatan Pendataan.xlsx'

        wb.save(response)
        return response



class MasterSurveyUploadClassView(LoginRequiredMixin, View):
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            form = forms.SurveiFormUpload(request.POST, request.FILES)
            model = models.SurveyModel
            
            if form.is_valid():
                df = form.cleaned_data
                objs = []
                for idx in range(len(df['id'])):
                    objs.append(
                        model(
                            nama = df['nama'][idx],
                            deskripsi= df['deskripsi'][idx],
                            tgl_mulai= df['tgl_mulai'][idx],
                            tgl_selesai= df['tgl_selesai'][idx],
                            state = df['state'][idx],
                        )
                    )
                model.objects.bulk_create(objs)
                return JsonResponse({"status": "success", "messages": f"<strong>{len(df['id'])}</strong> Data berhasil diupload."})
            else:
                error_messages = list(itertools.chain.from_iterable(form.errors['import_file'].as_data()))
                return JsonResponse({"status": "error", "messages": error_messages})

        return JsonResponse({"error": ""}, status=403)  


class MasterKegiatanSurveiClassView(LoginRequiredMixin, View):
    
    def get(self, request):
        context = {
            'title' : 'Kegiatan Survei',
            'form': forms.SubKegiatanSurveiForm()
        }
        
        return render(request, 'master_survey/kegiatan-survei.html', context)
    
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            pprint('Hello World Masuk')
            form = forms.SubKegiatanSurveiForm(request.POST)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil ditambahkan'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)
    
class MasterKegiatanSurveiDetailClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if is_ajax:
            if request.method == 'POST':
                id = request.POST.get('id')
                data_petugas = models.SubKegiatanSurvei.objects.filter(pk=id)
                
                if data_petugas.exists():
                    return JsonResponse({'status' : 'success', 'instance': list(data_petugas.values())[0]}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400) 

class MasterKegiatanSurveiUpdateClassView(LoginRequiredMixin, View):
    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            data = get_object_or_404(models.SubKegiatanSurvei, pk=request.POST.get('id'))
            form = forms.SubKegiatanSurveiForm(request.POST, instance=data)
            if form.is_valid():
                instance = form.save()
                ser_instance = serializers.serialize('json', [ instance, ])
                return JsonResponse({"instance": ser_instance, 'message': 'Data berhasil diubah'}, status=200)
            else:
                return JsonResponse({"error": form.errors}, status=400)
        return JsonResponse({"error": ""}, status=400)


class MasterKegiatanSurveiDeleteClassView(LoginRequiredMixin, View):

    def post(self, request):
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            if request.method == 'POST':
                pk = request.POST.get('id')
                data_kegiatan = models.SubKegiatanSurvei.objects.filter(pk = pk)
                if data_kegiatan.exists():
                    check_alokasi_mitra = AlokasiPetugas.objects.filter(sub_kegiatan = pk)
                    if check_alokasi_mitra.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada alokasi petugas.'}, status=200)
                    
                    check_kegiatan_penilaian = KegiatanPenilaianModel.objects.filter(kegiatan_survey = pk)
                    if check_kegiatan_penilaian.exists():
                        return JsonResponse({'status' : 'failed', 'message': 'Data survei telah digunakan pada master data penilaian.'}, status=200)
                
                    data_kegiatan.delete()
                    return JsonResponse({'status' : 'success', 'message': 'Data berhasil dihapus'}, status=200)
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Data tidak tersedia'}, status=200)
                
        return JsonResponse({'status': 'Invalid request'}, status=400)


class MasterKegiatanSurveiJsonClassView(LoginRequiredMixin, View):

    def post(self, request):
        data_wilayah = self._datatables(request)
        return HttpResponse(json.dumps(data_wilayah, cls=DjangoJSONEncoder), content_type='application/json')
		
    def _datatables(self, request):
        datatables = request.POST
        draw = int(datatables.get('draw'))
        start = int(datatables.get('start'))
        length = int(datatables.get('length'))
        page_number = int(start / length + 1)
        search = datatables.get('search[value]')

        order_idx = int(datatables.get('order[0][column]'))
        order_dir = datatables.get('order[0][dir]')
        order_col = 'columns[' + str(order_idx) + '][data]'
        order_col_name = datatables.get(order_col)

        if (order_dir == "desc"):
            order_col_name =  str('-' + order_col_name)

        data_kegiatan = models.SubKegiatanSurvei.objects

        if datatables.get('state_filter'):
            data_kegiatan = data_kegiatan.filter(status = datatables.get('state_filter'))

        if search:
            data_kegiatan = data_kegiatan.filter(
                Q(nama_kegiatan__icontains=search)|Q(survey__nama__icontains=search)|Q(survey__tgl_mulai__icontains=search)
            )

        data_kegiatan = data_kegiatan.exclude(Q(nama_kegiatan=None)|Q(survey=None)|Q(status=None))
        records_total = data_kegiatan.count()
        records_filtered = records_total
        
        data_kegiatan = data_kegiatan.order_by(order_col_name)
        paginator = Paginator(data_kegiatan, length)

        try:
            object_list = paginator.page(page_number).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(1).object_list

        data = []

        for obj in object_list:
            if obj.status == '0':
                bg_class = 'bg-warning'
            elif obj.status == '1':
                bg_class = 'bg-primary'
            else:
                bg_class = 'bg-success'

            data.append(
            {
                'nama_kegiatan': obj.nama_kegiatan,
                'survey__nama': obj.survey.nama,
                'survey__tgl_mulai': obj.survey.tgl_mulai.strftime('%Y'),
                'status': f'<span class="badge {bg_class} p-1">{obj.get_status_display()}</span>',
                'aksi': f'<a href="javascript:void(0);" onclick="editSubKegiatan({obj.id})" class="action-icon"><i class="mdi mdi-square-edit-outline font-15"></i></a> <a href="javascript:void(0);" onclick="deleteSubKegiatan({obj.id});" class="action-icon"> <i class="mdi mdi-delete font-15"></i></a>'
            })

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

